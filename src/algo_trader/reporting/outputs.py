"""Derivative Excel and Matplotlib presentation artifacts for reports."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg", force=True)

import matplotlib.pyplot as plt  # noqa: E402
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from algo_trader.reporting.models import ProfitFactor, ReportBundle
from algo_trader.reporting.tables import report_tables

MONEY_FORMAT = "₹#,##0.00;[Red]-₹#,##0.00"
PERCENT_FORMAT = "0.00%"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")


def _excel_value(value: object) -> object:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value
    if isinstance(value, (tuple, list, dict, set, frozenset)):
        return str(value)
    return value


def _write_table(sheet: Worksheet, rows: list[dict[str, Any]]) -> None:
    if not rows:
        sheet["A1"] = "No records"
        sheet.freeze_panes = "A2"
        return
    headers = list(rows[0])
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        sheet.append([_excel_value(row[header]) for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, start=1):
        values = [str(row.get(header, "")) for row in rows[:100]]
        width = min(max([len(header), *(len(value) for value in values)]) + 2, 35)
        sheet.column_dimensions[get_column_letter(index)].width = width
        lowered = header.lower()
        money_tokens = (
            "capital",
            "pnl",
            "cost",
            "turnover",
            "price",
            "margin",
            "notional",
            "amount",
        )
        if any(token in lowered for token in money_tokens):
            for cell in sheet[get_column_letter(index)][1:]:
                if isinstance(cell.value, int | float):
                    cell.number_format = MONEY_FORMAT
        elif any(token in lowered for token in ("return", "rate", "pct", "percentage")):
            for cell in sheet[get_column_letter(index)][1:]:
                if isinstance(cell.value, int | float):
                    cell.number_format = PERCENT_FORMAT


def _pf_display(value: ProfitFactor) -> Decimal | str:
    if value.is_unbounded:
        return "UNBOUNDED"
    if value.is_undefined:
        return "UNDEFINED"
    return value.value if value.value is not None else "UNDEFINED"


def _dashboard(workbook: Workbook, report: ReportBundle) -> Worksheet:
    sheet = workbook.active
    sheet.title = "Dashboard"
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "NEW ALGO TRADER — BACKTEST REPORT"
    sheet["A1"].font = Font(size=16, bold=True, color="1F4E78")
    sheet["A2"] = "ACTUAL REALIZED PORTFOLIO PERFORMANCE"
    sheet["A2"].font = Font(bold=True)
    performance = report.performance
    metrics = [
        ("Starting Capital", performance.starting_capital, MONEY_FORMAT),
        ("Ending Capital", performance.ending_capital, MONEY_FORMAT),
        ("Net Profit", performance.net_profit, MONEY_FORMAT),
        ("Total Return", performance.total_return, PERCENT_FORMAT),
        ("CAGR", performance.cagr, PERCENT_FORMAT),
        ("Actual Trades", performance.actual_trade_count, "0"),
        ("Win Rate", performance.win_rate, PERCENT_FORMAT),
        ("Net Profit Factor", _pf_display(performance.net_profit_factor), "0.00"),
        ("Avg Net Return/Trade", performance.average_net_return_per_trade, PERCENT_FORMAT),
        ("Max Realized Drawdown %", performance.maximum_realized_drawdown_pct, PERCENT_FORMAT),
        ("Trades/Day", performance.actual_trades_per_day, "0.00"),
        ("Total Costs", performance.total_costs, MONEY_FORMAT),
        ("Capital Exhausted", report.provenance.capital_exhausted, "General"),
    ]
    for row_index, (label, value, number_format) in enumerate(metrics, start=4):
        sheet.cell(row_index, 1, label).font = Font(bold=True)
        cell = sheet.cell(row_index, 2, _excel_value(value))
        cell.number_format = number_format

    sheet["D2"] = "HARD QUANTITATIVE TARGETS"
    sheet["D2"].font = Font(bold=True)
    statuses = [
        ("CAGR >20%", report.acceptance.cagr_pass),
        ("Win Rate >50%", report.acceptance.win_rate_pass),
        ("PF >2", report.acceptance.profit_factor_pass),
        ("Avg Net Return >0.5%", report.acceptance.average_net_return_pass),
        ("All Hard Targets", report.acceptance.hard_quantitative_targets_pass),
        ("FREQUENCY IDEAL >=2/day", report.acceptance.frequency_target_met),
    ]
    for row_index, (label, passed) in enumerate(statuses, start=4):
        sheet.cell(row_index, 4, label).font = Font(bold=True)
        sheet.cell(row_index, 5, "PASS" if passed else "FAIL")
    sheet.conditional_formatting.add(
        "E4:E9", CellIsRule(operator="equal", formula=['"PASS"'], fill=PASS_FILL)
    )
    sheet.conditional_formatting.add(
        "E4:E9", CellIsRule(operator="equal", formula=['"FAIL"'], fill=FAIL_FILL)
    )

    sheet["G2"] = "REPRODUCIBILITY"
    sheet["G2"].font = Font(bold=True)
    provenance = [
        ("report_id", report.provenance.report_id),
        ("run_id", report.provenance.run_id),
        ("git_commit", report.provenance.git_commit),
        ("reporting_version", report.provenance.reporting_version),
        ("backtester_version", report.provenance.backtester_version),
        ("cost_policy_id", report.provenance.cost_policy_id),
    ]
    if report.provenance.research_scope_id is not None:
        provenance.extend(
            [
                ("research_scope_id", report.provenance.research_scope_id),
                ("plan_id", report.provenance.plan_id),
                ("window_id", report.provenance.window_id),
            ]
        )
    for row_index, (label, value) in enumerate(provenance, start=4):
        sheet.cell(row_index, 7, label).font = Font(bold=True)
        sheet.cell(row_index, 8, value)
    for column, width in {"A": 28, "B": 18, "D": 28, "E": 12, "G": 22, "H": 28}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A4"
    return sheet


def _chart_data_sheet(
    workbook: Workbook,
    report: ReportBundle,
) -> tuple[Worksheet, dict[str, tuple[int, int, int]]]:
    sheet = workbook.create_sheet("Chart Data")
    positions: dict[str, tuple[int, int, int]] = {}
    row = 1

    def add(name: str, headers: tuple[str, str], values: list[tuple[object, object]]) -> None:
        nonlocal row
        start = row
        sheet.cell(row, 1, headers[0])
        sheet.cell(row, 2, headers[1])
        for first, second in values:
            row += 1
            sheet.cell(row, 1, _excel_value(first))
            sheet.cell(row, 2, _excel_value(second))
        positions[name] = (start, start + len(values), len(values))
        row += 2

    add(
        "equity",
        ("Exit Timestamp", "Realized Capital"),
        [(point.timestamp, point.realized_capital) for point in report.equity_curve],
    )
    add(
        "drawdown",
        ("Exit Timestamp", "Realized Drawdown"),
        [(point.timestamp, point.drawdown_pct) for point in report.equity_curve],
    )
    add(
        "daily",
        ("Trading Date", "Daily Net P&L"),
        [(point.trading_date, point.net_pnl) for point in report.daily_performance],
    )
    request = report.request_outcomes
    add(
        "requests",
        ("Request Outcome", "Count"),
        [
            ("Completed Actual", request.completed_actual),
            ("Completed Shadow", request.completed_shadow),
            ("Allocated No Fill", request.allocated_entry_not_filled),
            ("Shadow No Fill", request.shadow_entry_not_filled),
            ("Capital Exhausted", request.capital_exhausted),
        ],
    )
    add(
        "costs",
        ("Actual Cost Component", "Amount"),
        [
            (name, getattr(report.actual_costs, name))
            for name in report.actual_costs.component_percentages
        ],
    )
    add(
        "strategies",
        ("Strategy / Version", "Actual Net P&L"),
        [
            (f"{value.strategy_id} / {value.strategy_version}", value.net_pnl)
            for value in report.actual_strategy_breakdown
        ],
    )
    add(
        "exits",
        ("Exit Reason", "Actual Trade Count"),
        [
            (value.exit_reason.value, value.trade_count)
            for value in report.actual_exit_reason_breakdown
        ],
    )
    add(
        "symbols",
        ("Symbol", "Actual Net P&L"),
        [(value.symbol, value.net_pnl) for value in report.symbol_breakdown],
    )
    tables = report_tables(report)
    add(
        "monthly",
        ("Month", "Monthly Net P&L"),
        [
            (row["month"], row["net_pnl"])
            for row in tables["monthly_performance"].to_dicts()
        ],
    )
    add(
        "sides",
        ("Side", "Net P&L"),
        [
            (row["side"], row["net_pnl"])
            for row in tables["side_performance"].to_dicts()
        ],
    )
    sheet.sheet_state = "hidden"
    return sheet, positions


def _add_chart(
    dashboard: Worksheet,
    source: Worksheet,
    bounds: tuple[int, int, int],
    title: str,
    anchor: str,
    *,
    line: bool,
    percent: bool = False,
) -> None:
    start, end, count = bounds
    if count == 0:
        return
    chart = LineChart() if line else BarChart()
    chart.title = title
    chart.height = 7
    chart.width = 12
    chart.y_axis.title = "Percent" if percent else "Value"
    chart.x_axis.title = source.cell(start, 1).value
    chart.add_data(Reference(source, min_col=2, min_row=start, max_row=end), titles_from_data=True)
    chart.set_categories(Reference(source, min_col=1, min_row=start + 1, max_row=end))
    if not line:
        chart.type = "col"
    dashboard.add_chart(chart, anchor)


def _add_table_chart(
    sheet: Worksheet,
    title: str,
    anchor: str,
    *,
    category_header: str,
    value_headers: tuple[str, ...],
    line: bool = False,
    percent: bool = False,
) -> None:
    if sheet.max_row <= 1 or sheet["A1"].value == "No records":
        return
    headers = {cell.value: cell.column for cell in sheet[1]}
    chart = LineChart() if line else BarChart()
    chart.title = title
    chart.height = 7
    chart.width = 12
    chart.y_axis.title = "Percent" if percent else "Value"
    chart.x_axis.title = category_header
    for header in value_headers:
        column = headers[header]
        chart.add_data(
            Reference(sheet, min_col=column, min_row=1, max_row=sheet.max_row),
            titles_from_data=True,
        )
    category_column = headers[category_header]
    chart.set_categories(
        Reference(sheet, min_col=category_column, min_row=2, max_row=sheet.max_row)
    )
    if not line:
        chart.type = "col"
    sheet.add_chart(chart, anchor)


def _add_scatter_chart(
    sheet: Worksheet,
    title: str,
    anchor: str,
    *,
    x_header: str,
    y_header: str,
) -> None:
    if sheet.max_row <= 1 or sheet["A1"].value == "No records":
        return
    headers = {cell.value: cell.column for cell in sheet[1]}
    chart = ScatterChart()
    chart.title = title
    chart.height = 7
    chart.width = 12
    chart.x_axis.title = x_header
    chart.y_axis.title = y_header
    x_values = Reference(
        sheet, min_col=headers[x_header], min_row=2, max_row=sheet.max_row
    )
    y_values = Reference(
        sheet, min_col=headers[y_header], min_row=2, max_row=sheet.max_row
    )
    chart.series.append(Series(y_values, x_values, title_from_data=False))
    sheet.add_chart(chart, anchor)


def write_excel_report(report: ReportBundle, output_path: Path) -> Path:
    """Write a professional derivative workbook from canonical report tables."""
    if not isinstance(report, ReportBundle):
        raise TypeError("report must be a ReportBundle")
    path = Path(output_path)
    if path.exists():
        raise FileExistsError(f"Excel report already exists: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    tables = report_tables(report)
    workbook = Workbook()
    dashboard = _dashboard(workbook, report)
    sheet_tables = (
        ("Summary", "summary"),
        ("Actual Trades", "actual_trades"),
        ("Shadow Trades", "shadow_trades"),
        ("Daily Performance", "daily_performance"),
        ("Requests", "request_outcomes"),
        ("Equity Curve", "equity_curve"),
        ("Strategy Breakdown", "actual_strategy_breakdown"),
        ("Shadow Strategy", "shadow_strategy_breakdown"),
        ("Symbol Breakdown", "symbol_breakdown"),
        ("Costs", "actual_cost_breakdown"),
        ("Shadow Costs", "shadow_cost_breakdown"),
        ("Exit Reasons", "actual_exit_reason_breakdown"),
        ("Shadow Exit Reasons", "shadow_exit_reason_breakdown"),
        ("Provenance", "provenance"),
        ("Cumulative PnL", "cumulative_pnl"),
        ("Monthly Performance", "monthly_performance"),
        ("Side Performance", "side_performance"),
        ("Time of Day", "time_of_day_performance"),
        ("Holding Distribution", "holding_time_distribution"),
        ("Rolling 20 Trades", "rolling_trade_metrics"),
        ("Trade Diagnostics", "trade_diagnostics"),
        ("Cost Impact", "cost_impact"),
        ("Outcome Funnel", "outcome_funnel"),
        ("Actual vs Shadow", "actual_shadow_comparison"),
    )
    for sheet_name, table_name in sheet_tables:
        _write_table(workbook.create_sheet(sheet_name), tables[table_name].to_dicts())
    chart_data, positions = _chart_data_sheet(workbook, report)
    _add_chart(
        dashboard,
        chart_data,
        positions["equity"],
        "Realized Portfolio Equity",
        "A20",
        line=True,
    )
    _add_chart(
        dashboard,
        chart_data,
        positions["drawdown"],
        "Realized Capital Drawdown",
        "N20",
        line=True,
        percent=True,
    )
    _add_chart(dashboard, chart_data, positions["daily"], "Daily Actual Net P&L", "A35", line=False)
    _add_chart(
        dashboard,
        chart_data,
        positions["requests"],
        "Request Outcome Funnel",
        "N35",
        line=False,
    )
    _add_chart(
        dashboard,
        chart_data,
        positions["monthly"],
        "Monthly Actual Net P&L",
        "A65",
        line=False,
    )
    _add_chart(
        dashboard,
        chart_data,
        positions["sides"],
        "LONG vs SHORT Net P&L",
        "N65",
        line=False,
    )
    if report.actual_costs.total_costs > 0:
        _add_chart(
            dashboard,
            chart_data,
            positions["costs"],
            "Actual Transaction Cost Composition",
            "A50",
            line=False,
        )
    _add_chart(
        dashboard,
        chart_data,
        positions["strategies"],
        "Actual Strategy Net P&L",
        "N50",
        line=False,
    )
    _add_chart(
        workbook["Exit Reasons"],
        chart_data,
        positions["exits"],
        "Actual Exit Reason Counts",
        "H2",
        line=False,
    )
    _add_chart(
        dashboard,
        chart_data,
        positions["exits"],
        "Actual Exit Reason Mix",
        "A80",
        line=False,
    )
    if len(report.symbol_breakdown) <= 20:
        _add_chart(
            workbook["Symbol Breakdown"],
            chart_data,
            positions["symbols"],
            "Actual Symbol Net P&L",
            "Q2",
            line=False,
        )
    _add_table_chart(
        workbook["Cumulative PnL"],
        "Cumulative Gross vs Net P&L",
        "G2",
        category_header="trade_number",
        value_headers=("cumulative_gross_pnl", "cumulative_net_pnl"),
        line=True,
    )
    _add_table_chart(
        workbook["Cumulative PnL"],
        "Cumulative Cost Drag",
        "G17",
        category_header="trade_number",
        value_headers=("cumulative_cost_drag",),
        line=True,
    )
    for title, header, anchor, percent in (
        ("Monthly Net P&L", "net_pnl", "K2", False),
        ("Monthly Trade Count", "trade_count", "K17", False),
        ("Monthly Win Rate", "win_rate", "K32", True),
        ("Monthly Profit Factor", "profit_factor", "K47", False),
    ):
        _add_table_chart(
            workbook["Monthly Performance"],
            title,
            anchor,
            category_header="month",
            value_headers=(header,),
            percent=percent,
        )
    _add_table_chart(
        workbook["Side Performance"],
        "LONG vs SHORT Net P&L",
        "K2",
        category_header="side",
        value_headers=("net_pnl",),
    )
    _add_table_chart(
        workbook["Time of Day"],
        "30-Minute IST Entry-Time Performance",
        "K2",
        category_header="entry_time_bucket_ist",
        value_headers=("net_pnl",),
    )
    _add_table_chart(
        workbook["Holding Distribution"],
        "Holding-Time Distribution",
        "E2",
        category_header="bucket",
        value_headers=("trade_count",),
    )
    for title, header, anchor, percent in (
        ("Rolling 20-Trade Win Rate", "rolling_win_rate", "H2", True),
        ("Rolling 20-Trade Profit Factor", "profit_factor", "H17", False),
        (
            "Rolling 20-Trade Average Net Return",
            "rolling_average_net_return",
            "H32",
            True,
        ),
    ):
        _add_table_chart(
            workbook["Rolling 20 Trades"],
            title,
            anchor,
            category_header="trade_number",
            value_headers=(header,),
            line=True,
            percent=percent,
        )
    _add_scatter_chart(
        workbook["Trade Diagnostics"],
        "MFE vs MAE",
        "M2",
        x_header="mfe_return",
        y_header="mae_return",
    )
    _add_scatter_chart(
        workbook["Trade Diagnostics"],
        "MFE vs Realized Net Return",
        "M17",
        x_header="mfe_return",
        y_header="net_return",
    )
    _add_scatter_chart(
        workbook["Trade Diagnostics"],
        "MAE vs Realized Net Return",
        "M32",
        x_header="mae_return",
        y_header="net_return",
    )
    _add_table_chart(
        workbook["Cost Impact"],
        "Gross-to-Net Cost Impact",
        "E2",
        category_header="component",
        value_headers=("amount",),
    )
    _add_table_chart(
        workbook["Outcome Funnel"],
        "Request / Outcome Funnel",
        "E2",
        category_header="stage",
        value_headers=("count",),
    )
    _add_table_chart(
        workbook["Actual vs Shadow"],
        "Actual vs Shadow Net P&L",
        "J2",
        category_header="economic_status",
        value_headers=("net_pnl",),
    )
    workbook["Actual vs Shadow"]["J17"] = (
        "SHADOW IS HYPOTHETICAL — NO ACTUAL CAPITAL IMPACT"
    )
    workbook["Actual vs Shadow"]["J17"].font = Font(bold=True, color="C00000")
    workbook.save(path)
    return path


def _save_plot(path: Path, title: str, draw: Any) -> None:
    if path.exists():
        raise FileExistsError(f"visual report path already exists: {path}")
    figure, axis = plt.subplots(figsize=(10, 5), dpi=120)
    try:
        draw(axis)
        axis.set_title(title)
        figure.tight_layout()
        figure.savefig(path)
    finally:
        plt.close(figure)


def _empty(axis: Any, message: str = "No completed actual trades") -> None:
    axis.text(0.5, 0.5, message, ha="center", va="center", transform=axis.transAxes)
    axis.set_xticks([])
    axis.set_yticks([])


def write_visual_report(report: ReportBundle, output_directory: Path) -> tuple[Path, ...]:
    """Write deterministic headless PNG views from the same canonical tables."""
    if not isinstance(report, ReportBundle):
        raise TypeError("report must be a ReportBundle")
    directory = Path(output_directory)
    directory.mkdir(parents=True, exist_ok=True)
    filenames = [
        "realized_equity_curve.png",
        "realized_drawdown.png",
        "daily_net_pnl.png",
        "trade_net_returns.png",
        "strategy_net_pnl.png",
        "cost_composition.png",
        "request_outcomes.png",
        "exit_reason_counts.png",
        "cumulative_gross_vs_net_pnl.png",
        "cumulative_cost_drag.png",
        "monthly_net_pnl.png",
        "monthly_trade_count.png",
        "monthly_win_rate.png",
        "monthly_profit_factor.png",
        "long_short_performance.png",
        "time_of_day_performance.png",
        "holding_time_distribution.png",
        "mfe_mae_scatter.png",
        "mfe_vs_realized_return.png",
        "mae_vs_realized_return.png",
        "rolling_win_rate.png",
        "rolling_profit_factor.png",
        "rolling_average_net_return.png",
        "symbol_performance_top_bottom.png",
        "actual_vs_shadow_comparison.png",
    ]
    if report.shadow_trade_records:
        filenames.append("shadow_trade_net_returns.png")
    paths_to_write = tuple(directory / filename for filename in filenames)
    existing = tuple(path for path in paths_to_write if path.exists())
    if existing:
        raise FileExistsError(f"visual report path already exists: {existing[0]}")
    tables = report_tables(report)
    paths: list[Path] = []

    def save(filename: str, title: str, draw: Any) -> None:
        path = directory / filename
        _save_plot(path, title, draw)
        paths.append(path)

    equity = tables["equity_curve"].to_dicts()
    save(
        "realized_equity_curve.png",
        "Realized Portfolio Equity (Actual Only)",
        lambda axis: (
            axis.plot(
                [row["timestamp"] for row in equity],
                [float(row["realized_capital"]) for row in equity],
            ),
            axis.set_xlabel("Exit timestamp"),
            axis.set_ylabel("Realized capital (INR)"),
            axis.grid(alpha=0.25),
        ),
    )
    save(
        "realized_drawdown.png",
        "Realized Capital Drawdown (Actual Only)",
        lambda axis: (
            axis.plot(
                [row["timestamp"] for row in equity],
                [float(row["drawdown_pct"]) * 100 for row in equity],
            ),
            axis.set_xlabel("Exit timestamp"),
            axis.set_ylabel("Drawdown (%)"),
            axis.grid(alpha=0.25),
        ),
    )
    daily = tables["daily_performance"].to_dicts()
    save(
        "daily_net_pnl.png",
        "Daily Actual Net P&L",
        lambda axis: (
            axis.bar(
                [row["trading_date"].isoformat() for row in daily],
                [float(row["net_pnl"]) for row in daily],
            ),
            axis.set_xlabel("Evaluation trading date"),
            axis.set_ylabel("Net P&L (INR)"),
            axis.tick_params(axis="x", rotation=45),
            axis.axhline(0, color="black", linewidth=0.8),
        ),
    )
    trades = tables["actual_trades"].to_dicts()

    def trade_returns(axis: Any) -> None:
        if not trades:
            _empty(axis)
            return
        axis.plot(
            range(1, len(trades) + 1),
            [float(row["net_return"]) * 100 for row in trades],
            marker="o",
        )
        axis.axhline(0, color="black", linewidth=0.8)
        axis.set_xlabel("Chronological actual trade number")
        axis.set_ylabel("Net return (%)")

    save("trade_net_returns.png", "Actual Trade Net Returns", trade_returns)
    strategies = tables["actual_strategy_breakdown"].to_dicts()

    def strategy_pnl(axis: Any) -> None:
        if not strategies:
            _empty(axis)
            return
        axis.bar(
            [f"{row['strategy_id']} / {row['strategy_version']}" for row in strategies],
            [float(row["net_pnl"]) for row in strategies],
        )
        axis.axhline(0, color="black", linewidth=0.8)
        axis.set_ylabel("Actual net P&L (INR)")
        axis.tick_params(axis="x", rotation=30)

    save("strategy_net_pnl.png", "Actual Strategy Net P&L", strategy_pnl)
    costs = [row for row in tables["actual_cost_breakdown"].to_dicts() if row["is_cost_component"]]

    def cost_composition(axis: Any) -> None:
        if not costs or all(row["amount"] == 0 for row in costs):
            _empty(axis, "No actual transaction costs")
            return
        axis.bar([row["component"] for row in costs], [float(row["amount"]) for row in costs])
        axis.set_ylabel("Actual cost (INR)")
        axis.tick_params(axis="x", rotation=30)

    save("cost_composition.png", "Actual Transaction Cost Composition", cost_composition)
    request = report.request_outcomes
    request_values = [
        ("Completed Actual", request.completed_actual),
        ("Completed Shadow", request.completed_shadow),
        ("Allocated No Fill", request.allocated_entry_not_filled),
        ("Shadow No Fill", request.shadow_entry_not_filled),
        ("Capital Exhausted", request.capital_exhausted),
    ]
    save(
        "request_outcomes.png",
        "Request Outcome Funnel",
        lambda axis: (
            axis.bar([name for name, _ in request_values], [value for _, value in request_values]),
            axis.set_ylabel("Request count"),
            axis.tick_params(axis="x", rotation=30),
        ),
    )
    exits = tables["actual_exit_reason_breakdown"].to_dicts()
    save(
        "exit_reason_counts.png",
        "Actual Exit Reason Counts",
        lambda axis: (
            axis.bar([row["exit_reason"] for row in exits], [row["trade_count"] for row in exits]),
            axis.set_ylabel("Actual trade count"),
            axis.tick_params(axis="x", rotation=30),
        ),
    )
    cumulative = tables["cumulative_pnl"].to_dicts()

    def cumulative_gross_net(axis: Any) -> None:
        if not cumulative:
            _empty(axis)
            return
        numbers = [row["trade_number"] for row in cumulative]
        axis.plot(
            numbers,
            [float(row["cumulative_gross_pnl"]) for row in cumulative],
            label="Gross P&L",
        )
        axis.plot(
            numbers,
            [float(row["cumulative_net_pnl"]) for row in cumulative],
            label="Net P&L",
        )
        axis.set_xlabel("Completed actual trade number")
        axis.set_ylabel("Cumulative P&L (INR)")
        axis.axhline(0, color="black", linewidth=0.8)
        axis.legend()

    save(
        "cumulative_gross_vs_net_pnl.png",
        "Cumulative Gross vs Net P&L (Actual Only)",
        cumulative_gross_net,
    )

    def cumulative_cost(axis: Any) -> None:
        if not cumulative:
            _empty(axis)
            return
        axis.plot(
            [row["trade_number"] for row in cumulative],
            [float(row["cumulative_cost_drag"]) for row in cumulative],
        )
        axis.set_xlabel("Completed actual trade number")
        axis.set_ylabel("Cumulative costs (INR)")

    save("cumulative_cost_drag.png", "Cumulative Actual Cost Drag", cumulative_cost)
    monthly = tables["monthly_performance"].to_dicts()

    def monthly_bar(axis: Any, field: str, ylabel: str, *, percent: bool = False) -> None:
        if not monthly:
            _empty(axis)
            return
        scale = 100 if percent else 1
        axis.bar(
            [row["month"] for row in monthly],
            [
                float(row[field]) * scale
                if row[field] is not None
                else float("nan")
                for row in monthly
            ],
        )
        axis.set_xlabel("IST month (YYYY-MM)")
        axis.set_ylabel(ylabel)
        axis.tick_params(axis="x", rotation=45)

    save(
        "monthly_net_pnl.png",
        "Monthly Actual Net P&L",
        lambda axis: monthly_bar(axis, "net_pnl", "Net P&L (INR)"),
    )
    save(
        "monthly_trade_count.png",
        "Monthly Actual Trade Count",
        lambda axis: monthly_bar(axis, "trade_count", "Completed actual trades"),
    )
    save(
        "monthly_win_rate.png",
        "Monthly Actual Win Rate",
        lambda axis: monthly_bar(axis, "win_rate", "Win rate (%)", percent=True),
    )
    save(
        "monthly_profit_factor.png",
        "Monthly Actual Net Profit Factor",
        lambda axis: monthly_bar(axis, "profit_factor", "Profit factor"),
    )
    sides = tables["side_performance"].to_dicts()

    def side_performance(axis: Any) -> None:
        if not sides or not any(row["trade_count"] for row in sides):
            _empty(axis)
            return
        axis.bar(
            [row["side"] for row in sides],
            [float(row["net_pnl"]) for row in sides],
        )
        axis.axhline(0, color="black", linewidth=0.8)
        axis.set_ylabel("Actual net P&L (INR)")

    save(
        "long_short_performance.png",
        "LONG vs SHORT Actual Performance",
        side_performance,
    )
    time_rows = tables["time_of_day_performance"].to_dicts()

    def time_performance(axis: Any) -> None:
        if not time_rows:
            _empty(axis)
            return
        axis.bar(
            [row["entry_time_bucket_ist"] for row in time_rows],
            [float(row["net_pnl"]) for row in time_rows],
        )
        axis.axhline(0, color="black", linewidth=0.8)
        axis.set_xlabel("Actual entry-time bucket (IST, 30 minutes)")
        axis.set_ylabel("Net P&L (INR)")

    save(
        "time_of_day_performance.png",
        "30-Minute IST Entry-Time Performance",
        time_performance,
    )
    holding = tables["holding_time_distribution"].to_dicts()

    def holding_distribution(axis: Any) -> None:
        if not holding or not any(row["trade_count"] for row in holding):
            _empty(axis)
            return
        axis.bar(
            [row["bucket"] for row in holding],
            [row["trade_count"] for row in holding],
        )
        axis.set_xlabel("Holding time (minutes)")
        axis.set_ylabel("Completed actual trades")

    save(
        "holding_time_distribution.png",
        "Actual Holding-Time Distribution",
        holding_distribution,
    )
    diagnostics = tables["trade_diagnostics"].to_dicts()

    def scatter(axis: Any, x: str, y: str, x_label: str, y_label: str) -> None:
        if not diagnostics:
            _empty(axis)
            return
        axis.scatter(
            [float(row[x]) * 100 for row in diagnostics],
            [float(row[y]) * 100 for row in diagnostics],
        )
        axis.axhline(0, color="black", linewidth=0.8)
        axis.set_xlabel(x_label)
        axis.set_ylabel(y_label)

    save(
        "mfe_mae_scatter.png",
        "Trade-Level MFE vs MAE",
        lambda axis: scatter(axis, "mfe_return", "mae_return", "MFE (%)", "MAE (%)"),
    )
    save(
        "mfe_vs_realized_return.png",
        "MFE vs Realized Net Return",
        lambda axis: scatter(
            axis, "mfe_return", "net_return", "MFE (%)", "Realized net return (%)"
        ),
    )
    save(
        "mae_vs_realized_return.png",
        "MAE vs Realized Net Return",
        lambda axis: scatter(
            axis, "mae_return", "net_return", "MAE (%)", "Realized net return (%)"
        ),
    )
    rolling = tables["rolling_trade_metrics"].to_dicts()

    def rolling_line(axis: Any, field: str, ylabel: str, *, percent: bool = False) -> None:
        if not rolling:
            _empty(axis, "Fewer than 20 completed actual trades")
            return
        scale = 100 if percent else 1
        values = [
            float(row[field]) * scale if row[field] is not None else float("nan")
            for row in rolling
        ]
        axis.plot([row["trade_number"] for row in rolling], values, marker="o")
        axis.set_xlabel("Completed actual trade number")
        axis.set_ylabel(ylabel)

    save(
        "rolling_win_rate.png",
        "Rolling 20-Trade Win Rate",
        lambda axis: rolling_line(axis, "rolling_win_rate", "Win rate (%)", percent=True),
    )
    save(
        "rolling_profit_factor.png",
        "Rolling 20-Trade Net Profit Factor",
        lambda axis: rolling_line(axis, "profit_factor", "Profit factor"),
    )
    save(
        "rolling_average_net_return.png",
        "Rolling 20-Trade Average Net Return",
        lambda axis: rolling_line(
            axis,
            "rolling_average_net_return",
            "Average net return (%)",
            percent=True,
        ),
    )
    symbols = sorted(
        tables["symbol_breakdown"].to_dicts(),
        key=lambda row: (row["net_pnl"], row["symbol"]),
    )

    def symbol_performance(axis: Any) -> None:
        if not symbols:
            _empty(axis)
            return
        selected = symbols[:5]
        for row in symbols[-5:]:
            if row not in selected:
                selected.append(row)
        axis.bar(
            [row["symbol"] for row in selected],
            [float(row["net_pnl"]) for row in selected],
        )
        axis.axhline(0, color="black", linewidth=0.8)
        axis.set_xlabel("Bottom / top symbols by net P&L")
        axis.set_ylabel("Actual net P&L (INR)")

    save(
        "symbol_performance_top_bottom.png",
        "Symbol Performance — Bottom and Top",
        symbol_performance,
    )
    actual_shadow = tables["actual_shadow_comparison"].to_dicts()

    def actual_shadow_comparison(axis: Any) -> None:
        if not any(row["trade_count"] for row in actual_shadow):
            _empty(axis, "No completed actual or shadow trades")
            return
        axis.bar(
            ["ACTUAL", "SHADOW (HYPOTHETICAL)"],
            [float(row["net_pnl"]) for row in actual_shadow],
        )
        axis.axhline(0, color="black", linewidth=0.8)
        axis.set_ylabel("Net P&L (INR); shadow has no capital impact")

    save(
        "actual_vs_shadow_comparison.png",
        "Actual vs Shadow — HYPOTHETICAL / NO ACTUAL CAPITAL IMPACT",
        actual_shadow_comparison,
    )
    if report.shadow_trade_records:
        shadow = tables["shadow_trades"].to_dicts()
        save(
            "shadow_trade_net_returns.png",
            "SHADOW / HYPOTHETICAL — NO ACTUAL CAPITAL IMPACT",
            lambda axis: (
                axis.plot(
                    range(1, len(shadow) + 1),
                    [float(row["net_return"]) * 100 for row in shadow],
                    marker="o",
                ),
                axis.axhline(0, color="black", linewidth=0.8),
                axis.set_xlabel("Chronological shadow trade number"),
                axis.set_ylabel("Hypothetical net return (%)"),
            ),
        )
    return tuple(paths)
