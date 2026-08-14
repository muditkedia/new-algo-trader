"""Read-only comparison reporting across development or governed OOS runs."""

from __future__ import annotations

from collections.abc import Iterable
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font

from algo_trader.domain import Side
from algo_trader.reporting.models import (
    ReportBundle,
    ReportComparisonBundle,
    ReportComparisonRow,
)
from algo_trader.reporting.outputs import _empty, _save_plot, _write_table


def build_report_comparison(
    reports: Iterable[ReportBundle],
) -> ReportComparisonBundle:
    """Build chronological per-window metrics without pooling run economics."""
    selected = tuple(reports)
    if not selected:
        raise ValueError("report comparison requires at least one report")
    if any(not isinstance(report, ReportBundle) for report in selected):
        raise TypeError("all comparison inputs must be ReportBundle instances")

    provenance = [report.provenance for report in selected]
    oos_flags = [
        all(
            value is not None
            for value in (
                item.research_scope_id,
                item.plan_id,
                item.window_id,
            )
        )
        for item in provenance
    ]
    has_any_oos_value = any(
        value is not None
        for item in provenance
        for value in (item.research_scope_id, item.plan_id, item.window_id)
    )
    if has_any_oos_value and not all(oos_flags):
        raise ValueError(
            "OOS comparisons cannot mix missing or partial OOS provenance"
        )
    is_oos = all(oos_flags)
    research_scope_id = None
    plan_id = None
    if is_oos:
        scopes = {item.research_scope_id for item in provenance}
        plans = {item.plan_id for item in provenance}
        window_ids = [item.window_id for item in provenance]
        if len(scopes) != 1:
            raise ValueError("OOS comparison requires one research_scope_id")
        if len(plans) != 1:
            raise ValueError("OOS comparison requires one plan_id")
        if len(window_ids) != len(set(window_ids)):
            raise ValueError("OOS comparison window_id values must be unique")
        research_scope_id = next(iter(scopes))
        plan_id = next(iter(plans))

    ordered = sorted(
        selected,
        key=lambda report: (
            report.provenance.window_start,
            report.provenance.window_end,
            report.provenance.window_id or "",
            report.provenance.report_id,
        ),
    )
    rows = []
    for report in ordered:
        long_count = sum(
            record.trade.signal.side is Side.LONG
            for record in report.actual_trade_records
        )
        short_count = sum(
            record.trade.signal.side is Side.SHORT
            for record in report.actual_trade_records
        )
        rows.append(
            ReportComparisonRow(
                report_id=report.provenance.report_id,
                run_id=report.provenance.run_id,
                backtester_version=report.provenance.backtester_version,
                reporting_version=report.provenance.reporting_version,
                research_scope_id=report.provenance.research_scope_id,
                plan_id=report.provenance.plan_id,
                window_id=report.provenance.window_id,
                window_start=report.provenance.window_start,
                window_end=report.provenance.window_end,
                actual_trade_count=report.performance.actual_trade_count,
                actual_trades_per_day=report.performance.actual_trades_per_day,
                net_pnl=report.performance.net_profit,
                ending_capital=report.performance.ending_capital,
                cagr=report.performance.cagr,
                win_rate=report.performance.win_rate,
                profit_factor=report.performance.net_profit_factor,
                average_net_return=report.performance.average_net_return_per_trade,
                max_drawdown=report.performance.maximum_realized_drawdown_pct,
                total_actual_costs=report.performance.total_costs,
                long_trade_count=long_count,
                short_trade_count=short_count,
                **report.acceptance.model_dump(mode="python"),
            )
        )
    return ReportComparisonBundle(
        is_oos=is_oos,
        research_scope_id=research_scope_id,
        plan_id=plan_id,
        rows=tuple(rows),
    )


def _comparison_rows(comparison: ReportComparisonBundle) -> list[dict[str, object]]:
    rows = []
    for row in comparison.rows:
        values = row.model_dump(mode="python", exclude={"profit_factor"})
        values.update(
            {
                "profit_factor": row.profit_factor.value,
                "profit_factor_is_unbounded": row.profit_factor.is_unbounded,
                "profit_factor_is_undefined": row.profit_factor.is_undefined,
            }
        )
        rows.append(values)
    return rows


def _add_comparison_chart(
    dashboard: Any,
    raw: Any,
    *,
    title: str,
    value_header: str,
    anchor: str,
    label_header: str,
    line: bool = False,
) -> None:
    headers = {cell.value: cell.column for cell in raw[1]}
    chart = LineChart() if line else BarChart()
    chart.title = title
    chart.height = 7
    chart.width = 12
    chart.add_data(
        Reference(
            raw,
            min_col=headers[value_header],
            min_row=1,
            max_row=raw.max_row,
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            raw,
            min_col=headers[label_header],
            min_row=2,
            max_row=raw.max_row,
        )
    )
    if not line:
        chart.type = "col"
    dashboard.add_chart(chart, anchor)


def write_comparison_excel_report(
    comparison: ReportComparisonBundle,
    output_path: Path,
) -> Path:
    """Write an auditable native-chart workbook for run/window comparison."""
    if not isinstance(comparison, ReportComparisonBundle):
        raise TypeError("comparison must be a ReportComparisonBundle")
    path = Path(output_path)
    if path.exists():
        raise FileExistsError(f"comparison Excel report already exists: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Comparison Dashboard"
    dashboard["A1"] = "NEW ALGO TRADER — RUN / OOS WINDOW COMPARISON"
    dashboard["A1"].font = Font(size=16, bold=True, color="1F4E78")
    dashboard["A3"] = "Research scope"
    dashboard["B3"] = comparison.research_scope_id or "DEVELOPMENT"
    dashboard["A4"] = "Plan"
    dashboard["B4"] = comparison.plan_id or "N/A"
    dashboard["A5"] = "Metrics remain per-window; no combined CAGR is calculated."
    raw = workbook.create_sheet("Raw Comparison")
    _write_table(raw, _comparison_rows(comparison))
    for index, (title, header) in enumerate(
        (
            ("Net P&L by Window", "net_pnl"),
            ("Win Rate by Window", "win_rate"),
            ("Profit Factor by Window", "profit_factor"),
            ("Average Net Return by Window", "average_net_return"),
            ("Max Drawdown by Window", "max_drawdown"),
            ("Trade Frequency by Window", "actual_trades_per_day"),
            ("Actual Costs by Window", "total_actual_costs"),
            ("Acceptance Stability", "hard_quantitative_targets_pass"),
        )
    ):
        row = 8 + (index // 2) * 15
        column = "A" if index % 2 == 0 else "N"
        _add_comparison_chart(
            dashboard,
            raw,
            title=title,
            value_header=header,
            anchor=f"{column}{row}",
            label_header="window_id" if comparison.is_oos else "report_id",
        )
    workbook.save(path)
    return path


COMPARISON_VISUAL_FILENAMES = (
    "window_net_pnl.png",
    "window_win_rate.png",
    "window_profit_factor.png",
    "window_average_net_return.png",
    "window_max_drawdown.png",
    "window_trade_frequency.png",
    "window_costs.png",
    "cumulative_window_net_pnl.png",
)


def write_comparison_visual_report(
    comparison: ReportComparisonBundle,
    output_directory: Path,
) -> tuple[Path, ...]:
    """Write headless per-window diagnostics without pooling window metrics."""
    if not isinstance(comparison, ReportComparisonBundle):
        raise TypeError("comparison must be a ReportComparisonBundle")
    directory = Path(output_directory)
    paths = tuple(directory / name for name in COMPARISON_VISUAL_FILENAMES)
    existing = next((path for path in paths if path.exists()), None)
    if existing is not None:
        raise FileExistsError(f"comparison visual already exists: {existing}")
    directory.mkdir(parents=True, exist_ok=True)
    labels = [row.window_id or row.report_id for row in comparison.rows]

    def save(filename: str, title: str, field: str, ylabel: str) -> None:
        def draw(axis: Any) -> None:
            values = []
            for row in comparison.rows:
                value = getattr(row, field)
                if field == "profit_factor":
                    value = row.profit_factor.value
                values.append(float(value) if value is not None else float("nan"))
            if all(value != value for value in values):
                _empty(axis, "Metric undefined for all windows")
                return
            axis.bar(labels, values)
            axis.set_xlabel("Window / report")
            axis.set_ylabel(ylabel)
            axis.tick_params(axis="x", rotation=45)

        _save_plot(directory / filename, title, draw)

    save("window_net_pnl.png", "Net P&L by Window", "net_pnl", "Net P&L (INR)")
    save("window_win_rate.png", "Win Rate by Window", "win_rate", "Win rate")
    save(
        "window_profit_factor.png",
        "Net Profit Factor by Window",
        "profit_factor",
        "Profit factor",
    )
    save(
        "window_average_net_return.png",
        "Average Net Return by Window",
        "average_net_return",
        "Average net return",
    )
    save(
        "window_max_drawdown.png",
        "Max Drawdown by Window",
        "max_drawdown",
        "Max realized drawdown",
    )
    save(
        "window_trade_frequency.png",
        "Trade Frequency by Window",
        "actual_trades_per_day",
        "Actual trades / evaluation day",
    )
    save(
        "window_costs.png",
        "Actual Costs by Window",
        "total_actual_costs",
        "Actual costs (INR)",
    )

    cumulative: list[Decimal] = []
    value = Decimal("0")
    for row in comparison.rows:
        value += row.net_pnl
        cumulative.append(value)

    def draw_cumulative(axis: Any) -> None:
        axis.plot(labels, [float(item) for item in cumulative], marker="o")
        axis.axhline(0, color="black", linewidth=0.8)
        axis.set_xlabel("Window / report")
        axis.set_ylabel("Cumulative window net P&L (INR)")
        axis.tick_params(axis="x", rotation=45)

    _save_plot(
        directory / "cumulative_window_net_pnl.png",
        "Cumulative Net P&L Across Ordered Windows (No Pooled CAGR)",
        draw_cumulative,
    )
    return paths
