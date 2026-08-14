"""Integrity-checked cumulative research artifacts and atomic run directories."""

from __future__ import annotations

import hashlib
import json
import os
import shutil
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from algo_trader.backtest import BacktestRunResult
from algo_trader.reporting import ReportBundle, report_tables
from algo_trader.research.models import ResearchDecisionRecord

REQUIRED_RUN_ARTIFACTS = (
    "backtest_result.json",
    "report_bundle.json",
    "run_manifest.json",
)
REVIEW_JSON_NAME = "strategy1_research_history.json"
REVIEW_WORKBOOK_NAME = "strategy1_research_master.xlsx"


class ResearchArtifactIntegrityError(RuntimeError):
    """Canonical research evidence is present but invalid or inconsistent."""


@dataclass(frozen=True, slots=True)
class CompletedResearchRun:
    directory: Path
    result: BacktestRunResult
    report: ReportBundle
    manifest: Mapping[str, Any]
    source_sha256: Mapping[str, str]


@dataclass(frozen=True, slots=True)
class ResearchRunDiscovery:
    completed: tuple[CompletedResearchRun, ...]
    skipped_incomplete: tuple[tuple[Path, str], ...]


def discover_research_runs(results_root: Path) -> ResearchRunDiscovery:
    """Load complete runs; skip explicit incompletes and fail on corrupt evidence."""
    runs_root = Path(results_root) / "runs"
    if not runs_root.exists():
        return ResearchRunDiscovery((), ())
    completed: list[CompletedResearchRun] = []
    skipped: list[tuple[Path, str]] = []
    for directory in sorted(path for path in runs_root.iterdir() if path.is_dir()):
        if directory.name.endswith(".tmp") or directory.name.startswith("_incomplete"):
            skipped.append((directory, "explicit staging/incomplete directory"))
            continue
        paths = {name: directory / name for name in REQUIRED_RUN_ARTIFACTS}
        missing = tuple(name for name, path in paths.items() if not path.is_file())
        if "run_manifest.json" in missing:
            skipped.append((directory, "completion manifest is missing"))
            continue
        if missing:
            raise ResearchArtifactIntegrityError(
                f"completed run {directory.name} is missing required artifacts: {missing}"
            )
        try:
            result = _load_backtest_result(paths["backtest_result.json"])
            report = _load_report_bundle(paths["report_bundle.json"])
            manifest = json.loads(paths["run_manifest.json"].read_text(encoding="utf-8"))
        except Exception as error:
            raise ResearchArtifactIntegrityError(
                f"completed run {directory.name} contains unreadable/corrupt evidence"
            ) from error
        if not isinstance(manifest, Mapping):
            raise ResearchArtifactIntegrityError(
                f"completed run {directory.name} manifest must be an object"
            )
        if report.provenance.run_id != result.run_id:
            raise ResearchArtifactIntegrityError(
                f"completed run {directory.name} has inconsistent run identity"
            )
        hashes = {name: _sha256(path) for name, path in paths.items()}
        declared = manifest.get("artifact_sha256")
        if declared is not None:
            if not isinstance(declared, Mapping) or any(
                declared.get(name) != digest for name, digest in hashes.items()
            ):
                raise ResearchArtifactIntegrityError(
                    f"completed run {directory.name} artifact fingerprint mismatch"
                )
        completed.append(
            CompletedResearchRun(directory, result, report, manifest, hashes)
        )
    completed.sort(
        key=lambda run: (
            str(run.manifest.get("completed_at", "")),
            run.result.run_id,
        )
    )
    return ResearchRunDiscovery(tuple(completed), tuple(skipped))


def prepare_staging_run_directory(runs_root: Path, run_id: str) -> tuple[Path, Path]:
    """Create a non-completed staging directory for one deterministic run ID."""
    root = Path(runs_root).resolve()
    final = root / run_id
    staging = root / f"{run_id}.tmp"
    if final.exists() or staging.exists():
        raise FileExistsError(f"research run path already exists for {run_id}")
    root.mkdir(parents=True, exist_ok=True)
    staging.mkdir()
    return staging, final


def finalize_staging_run_directory(staging: Path, final: Path) -> Path:
    """Validate canonical artifacts and atomically expose a completed run directory."""
    staging = Path(staging).resolve()
    final = Path(final).resolve()
    if staging.parent != final.parent or staging.name != f"{final.name}.tmp":
        raise ValueError("staging/final run paths do not form an atomic pair")
    missing = tuple(name for name in REQUIRED_RUN_ARTIFACTS if not (staging / name).is_file())
    if missing:
        raise ResearchArtifactIntegrityError(
            f"staging run is missing required artifacts: {missing}"
        )
    _validate_run_directory(staging)
    if final.exists():
        raise FileExistsError(f"completed research run already exists: {final}")
    os.replace(staging, final)
    return final


def build_review_artifacts(
    *,
    results_root: Path,
    research_scope_id: str,
    decisions: Iterable[ResearchDecisionRecord],
    oos_plan: object,
    generated_at: datetime,
) -> tuple[Path, Path]:
    """Atomically rebuild and validate the exact two-file review contract."""
    root = Path(results_root).resolve()
    discovery = discover_research_runs(root)
    if not discovery.completed:
        raise ResearchArtifactIntegrityError("no completed research runs are available")
    selected_decisions = tuple(decisions)
    if any(not isinstance(item, ResearchDecisionRecord) for item in selected_decisions):
        raise TypeError("decisions must contain ResearchDecisionRecord values")

    review = root / "UPLOAD_FOR_REVIEW"
    staging = root / "UPLOAD_FOR_REVIEW.tmp"
    backup = root / "UPLOAD_FOR_REVIEW.previous"
    if staging.exists() or backup.exists():
        raise FileExistsError("stale review staging/backup directory requires inspection")
    staging.mkdir(parents=True)
    json_path = staging / REVIEW_JSON_NAME
    workbook_path = staging / REVIEW_WORKBOOK_NAME
    try:
        _write_review_json(
            json_path,
            research_scope_id,
            discovery,
            selected_decisions,
            oos_plan,
            generated_at,
        )
        _write_review_workbook(workbook_path, discovery, selected_decisions, oos_plan)
        _validate_review_files(json_path, workbook_path)
    except Exception:
        shutil.rmtree(staging)
        raise
    if review.exists():
        os.replace(review, backup)
    try:
        os.replace(staging, review)
    except Exception:
        if backup.exists() and not review.exists():
            os.replace(backup, review)
        raise
    if backup.exists():
        shutil.rmtree(backup)
    final_json = review / REVIEW_JSON_NAME
    final_workbook = review / REVIEW_WORKBOOK_NAME
    _validate_review_files(final_json, final_workbook)
    if tuple(sorted(path.name for path in review.iterdir())) != (
        REVIEW_JSON_NAME,
        REVIEW_WORKBOOK_NAME,
    ):
        raise ResearchArtifactIntegrityError("review folder must contain exactly two files")
    return final_json, final_workbook


def _write_review_json(
    path: Path,
    scope: str,
    discovery: ResearchRunDiscovery,
    decisions: tuple[ResearchDecisionRecord, ...],
    oos_plan: object,
    generated_at: datetime,
) -> None:
    runs = []
    for run in discovery.completed:
        runs.append(
            {
                "run_id": run.result.run_id,
                "strategy_versions": run.result.strategy_versions,
                "phase": run.manifest.get("phase"),
                "window_id": run.report.provenance.window_id,
                "source_sha256": dict(run.source_sha256),
                "run_input_fingerprint": run.manifest.get("run_input_fingerprint"),
                "backtest_result": run.result.model_dump(mode="json"),
                "report_bundle": run.report.model_dump(mode="json"),
                "run_manifest": dict(run.manifest),
            }
        )
    payload = {
        "schema_version": "2",
        "research_scope_id": scope,
        "generated_at": generated_at.isoformat(),
        "runs": runs,
        "research_decisions": [item.model_dump(mode="json") for item in decisions],
        "oos_plan": (
            oos_plan.model_dump(mode="json")
            if hasattr(oos_plan, "model_dump")
            else oos_plan
        ),
        "skipped_incomplete_runs": [
            {"directory": path.name, "reason": reason}
            for path, reason in discovery.skipped_incomplete
        ],
    }
    path.write_text(
        json.dumps(payload, indent=2, sort_keys=True),
        encoding="utf-8",
        newline="\n",
    )


def _write_review_workbook(
    path: Path,
    discovery: ResearchRunDiscovery,
    decisions: tuple[ResearchDecisionRecord, ...],
    oos_plan: object,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _append_rows(
        workbook.create_sheet("Run History"),
        [
            {
                "run_id": run.result.run_id,
                "phase": run.manifest.get("phase"),
                "window_id": run.report.provenance.window_id,
                "window_start": run.result.window_start,
                "window_end": run.result.window_end,
                "actual_trades": len(run.result.actual_trade_records),
                "shadow_trades": len(run.result.shadow_trade_records),
            }
            for run in discovery.completed
        ],
    )
    _append_rows(
        workbook.create_sheet("Research Decisions"),
        [item.model_dump(mode="json") for item in decisions],
    )
    _append_rows(workbook.create_sheet("OOS State Plan"), [_object_row(oos_plan)])
    table_sheets = {
        "Dashboard": "summary",
        "Actual Trades": "actual_trades",
        "Shadow Trades": "shadow_trades",
        "Request Outcomes": "request_outcomes",
        "Equity Curve": "equity_curve",
        "Daily Performance": "daily_performance",
        "Monthly Performance": "monthly_performance",
        "Side Breakdown": "side_performance",
        "Symbol Breakdown": "symbol_breakdown",
        "Exit Reason Detail": "actual_exit_reason_breakdown",
        "Cost Breakdown": "actual_cost_breakdown",
        "MFE MAE R Diagnostics": "trade_diagnostics",
        "Time of Day": "time_of_day_performance",
        "Holding Distribution": "holding_time_distribution",
        "Rolling Metrics": "rolling_trade_metrics",
        "Cost Impact": "cost_impact",
        "Outcome Funnel": "outcome_funnel",
        "Actual vs Shadow": "actual_shadow_comparison",
        "Provenance": "provenance",
    }
    for sheet_name, table_name in table_sheets.items():
        rows: list[dict[str, object]] = []
        for run in discovery.completed:
            for row in report_tables(run.report)[table_name].to_dicts():
                rows.append({"run_id": run.result.run_id, **row})
        _append_rows(workbook.create_sheet(sheet_name), rows)
    _append_rows(workbook.create_sheet("Gate History"), _gate_rows(discovery.completed))
    _append_rows(
        workbook.create_sheet("Signal Features"),
        _signal_feature_rows(discovery.completed),
    )
    workbook.save(path)


def _signal_feature_rows(runs: tuple[CompletedResearchRun, ...]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for run in runs:
        records = (*run.result.actual_trade_records, *run.result.shadow_trade_records)
        for record in records:
            signal = record.trade.signal
            row: dict[str, object] = {
                "run_id": run.result.run_id,
                "candidate_identity": record.allocation_identity,
                "symbol": signal.symbol,
                "side": signal.side.value,
                "signal_timestamp": signal.timestamp,
            }
            row.update(
                {
                    f"parameter.{key}": value
                    for key, value in signal.strategy_parameters.items()
                }
            )
            row.update({f"feature.{key}": value for key, value in signal.feature_snapshot.items()})
            rows.append(row)
    return rows


def _gate_rows(runs: tuple[CompletedResearchRun, ...]) -> list[dict[str, object]]:
    return [
        {
            "run_id": run.result.run_id,
            **(
                dict(run.manifest.get("causality_gate", {}))
                if isinstance(run.manifest.get("causality_gate"), Mapping)
                else {}
            ),
        }
        for run in runs
    ]


def _append_rows(sheet: object, rows: list[dict[str, object]]) -> None:
    if not rows:
        sheet.append(["No records"])
        return
    headers = sorted({key for row in rows for key in row})
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([_excel_value(row.get(header)) for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _excel_value(value: object) -> object:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, (Mapping, tuple, list)):
        return json.dumps(value, sort_keys=True, default=str)
    return value


def _object_row(value: object) -> dict[str, object]:
    if value is None:
        return {"status": "OOS plan unavailable"}
    if hasattr(value, "model_dump"):
        return value.model_dump(mode="json")
    if isinstance(value, Mapping):
        return dict(value)
    return {"value": str(value)}


def _validate_run_directory(directory: Path) -> None:
    paths = {name: directory / name for name in REQUIRED_RUN_ARTIFACTS}
    try:
        result = _load_backtest_result(paths["backtest_result.json"])
        report = _load_report_bundle(paths["report_bundle.json"])
        manifest = json.loads(paths["run_manifest.json"].read_text(encoding="utf-8"))
    except Exception as error:
        raise ResearchArtifactIntegrityError(
            "staging run contains unreadable/corrupt canonical evidence"
        ) from error
    if not isinstance(manifest, Mapping):
        raise ResearchArtifactIntegrityError("staging run manifest must be an object")
    if report.provenance.run_id != result.run_id:
        raise ResearchArtifactIntegrityError("staging run has inconsistent run identity")


def _validate_review_files(json_path: Path, workbook_path: Path) -> None:
    if not json_path.is_file() or json_path.stat().st_size == 0:
        raise ResearchArtifactIntegrityError("cumulative JSON is missing or empty")
    try:
        parsed = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as error:
        raise ResearchArtifactIntegrityError("cumulative JSON cannot be parsed") from error
    if not isinstance(parsed, Mapping) or not parsed.get("runs"):
        raise ResearchArtifactIntegrityError("cumulative JSON has no complete run records")
    if not workbook_path.is_file() or workbook_path.stat().st_size == 0:
        raise ResearchArtifactIntegrityError("cumulative workbook is missing or empty")
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
        workbook.close()
    except Exception as error:
        raise ResearchArtifactIntegrityError("cumulative workbook cannot be reopened") from error


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _load_backtest_result(path: Path) -> BacktestRunResult:
    """Read current JSON plus legacy files that serialized computed cost totals.

    The compatibility transform is in-memory and removes only values that the
    frozen cost models deterministically recompute. Canonical source evidence is
    never rewritten and its SHA-256 therefore remains unchanged.
    """
    payload = json.loads(path.read_text(encoding="utf-8"))
    _remove_legacy_computed_cost_totals(payload)
    return BacktestRunResult.model_validate_json(
        json.dumps(payload, sort_keys=True, separators=(",", ":"))
    )


def _load_report_bundle(path: Path) -> ReportBundle:
    payload = json.loads(path.read_text(encoding="utf-8"))
    _remove_legacy_computed_cost_totals(payload)
    return ReportBundle.model_validate_json(
        json.dumps(payload, sort_keys=True, separators=(",", ":"))
    )


def _remove_legacy_computed_cost_totals(value: object) -> None:
    if isinstance(value, dict):
        costs = value.get("round_trip_cost_breakdown")
        if isinstance(costs, dict):
            costs.pop("total", None)
            for leg in ("entry", "exit"):
                breakdown = costs.get(leg)
                if isinstance(breakdown, dict):
                    breakdown.pop("total", None)
        for child in value.values():
            _remove_legacy_computed_cost_totals(child)
    elif isinstance(value, list):
        for child in value:
            _remove_legacy_computed_cost_totals(child)
