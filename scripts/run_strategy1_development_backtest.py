from __future__ import annotations

import argparse
import hashlib
import json
import os
import subprocess
from collections.abc import Iterable, Mapping
from dataclasses import asdict
from datetime import UTC, date, datetime, time
from decimal import ROUND_CEILING, Decimal
from pathlib import Path
from uuid import uuid4
from zoneinfo import ZoneInfo

import duckdb
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import polars as pl
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from algo_trader.backtest import (
    BACKTESTER_VERSION,
    BacktestConfig,
    BacktestTradeRequest,
    DynamicExitPolicySpec,
    HistoricalBacktester,
    RMultipleTrailingExitPolicyResolver,
)
from algo_trader.broker import (
    AngelOneBroker,
    AngelOneLiveMarginProvider,
    BrokerSystemicError,
    HistoricalMarginRequirementProvider,
    HistoricalMarginSnapshot,
    create_historical_margin_snapshot,
    create_margin_snapshot_entry,
    fetch_instrument_master,
    load_historical_margin_snapshot,
    save_historical_margin_snapshot,
)
from algo_trader.costs import BrokeragePlan
from algo_trader.data import (
    MarketDataConfig,
    ParquetMarketDataStore,
    SymbolCoverage,
    bar_available_at,
)
from algo_trader.domain import (
    MAX_NOTIONAL,
    MIN_NOTIONAL,
    NOTIONAL_INCREMENT,
    MLScore,
    OrderIntent,
    OrderType,
    Side,
    Signal,
)
from algo_trader.execution import FixedBasisPointsSlippage, HistoricalExecutionSimulator
from algo_trader.ml import MetaFeatureSchema, TradeScorer
from algo_trader.oos import (
    STANDARD_OOS_PARTITION_POLICY,
    OOSAuditContext,
    OOSPlan,
    OOSRegistry,
    OOSWindowState,
    create_standard_oos_plan,
    derive_equity_data_horizon,
    fingerprint_backtest_result,
    select_historically_available_equities,
    shift_calendar_months,
)
from algo_trader.portfolio import AllocationCandidate, MarginRequirementQuote, PortfolioState
from algo_trader.reporting import (
    REPORTING_VERSION,
    ReportBundle,
    ReportContext,
    build_report,
    report_tables,
    write_report_dataset,
    write_visual_report,
)
from algo_trader.research import (
    ResearchDecisionRecord,
    build_market_data_manifest,
    build_review_artifacts,
    canonical_fingerprint,
    discover_research_runs,
    environment_snapshot,
    finalize_staging_run_directory,
    prepare_staging_run_directory,
    score_and_build_requests,
    select_scope_trade_scorer,
)
from algo_trader.runtime import load_smartapi_credentials
from algo_trader.strategies import assert_strategy_prefix_invariant
from algo_trader.strategies.causality import StrategyCausalityReport
from algo_trader.strategies.liquidity_shock_reclaim import LiquidityShockReclaimStrategy

IST = ZoneInfo("Asia/Kolkata")

ROOT = Path(__file__).resolve().parents[1]
RUNNER_PATH = Path(__file__).resolve()
RUNNER_RELATIVE_PATH = RUNNER_PATH.relative_to(ROOT).as_posix()

DATA_PATH = ROOT / "data" / "market" / "NSE" / "5M"
RESULTS_ROOT = (
    ROOT
    / "results"
    / "research"
    / "liquidity-shock-exhaustion-reclaim"
)
OOS_REGISTRY_PATH = RESULTS_ROOT / "oos_registry.duckdb"
UPLOAD_REVIEW_DIR = RESULTS_ROOT / "UPLOAD_FOR_REVIEW"
MASTER_EXCEL_PATH = UPLOAD_REVIEW_DIR / "strategy1_research_master.xlsx"
MASTER_JSON_PATH = UPLOAD_REVIEW_DIR / "strategy1_research_history.json"
MASTER_VISUAL_DIR = RESULTS_ROOT / "research_visuals"
MARGIN_SNAPSHOT_DIR = RESULTS_ROOT / "margin_snapshots"
MARKET_DATA_FINGERPRINT_CACHE = RESULTS_ROOT / "market_data_fingerprint_cache.json"
MODEL_ROOT = RESULTS_ROOT / "models"
HISTORICAL_SYMBOL_ALIAS_PATH = RESULTS_ROOT / "historical_symbol_aliases.json"

RESEARCH_SCOPE_ID = "liquidity-shock-exhaustion-reclaim"
PLAN_ID = "liquidity-shock-exhaustion-reclaim-standard-v1"

INITIAL_CAPITAL = Decimal("100000")
# Explicit research assumption. This is intentionally not a strategy parameter.
DEFAULT_SLIPPAGE_BPS = Decimal("5")
DEFAULT_BROKERAGE_PLAN = BrokeragePlan.PLUS

# Development begins at the earliest equity data date. Strategy warm-up consumes the earliest
# portion. The first scored in-sample development evaluation begins at the earliest causal
# Strategy-1-ready date and lasts exactly three calendar months. OOS begins immediately after it.
DEVELOPMENT_EVALUATION_MONTHS = 3

MARGIN_REFERENCE_NOTIONAL = 100_000
FORCED_BACKTEST_EXIT_TIME = time(15, 25)

BOOTSTRAP_MODEL_VERSION = "bootstrap-neutral-v1"
DYNAMIC_EXIT_POLICY_ID = "R_MULTIPLE_TRAILING_V1"
STRATEGY1_META_FEATURE_SCHEMA = MetaFeatureSchema(
    feature_names=(
        "shock_return",
        "shock_robust_z",
        "relative_volume",
        "median_daily_turnover",
        "atr",
        "penetration_atr",
        "reclaim_atr",
        "confirmation_return_from_event_close",
    )
)

# Human decisions link to canonical artifacts; numerical metrics remain owned by those artifacts.
RESEARCH_DECISION_HISTORY = (
    ResearchDecisionRecord(
        decision_id="strategy1-v1.0-development-review",
        source_run_ids=(
            "strategy1-development-2016-12-29-to-2017-03-29-1.0.0-a289fe2e087f",
        ),
        strategy_id=RESEARCH_SCOPE_ID,
        strategy_version="1.0.0",
        research_scope_id=RESEARCH_SCOPE_ID,
        decision="ITERATE_IN_DEVELOPMENT",
        diagnosis="The broad baseline had negative gross expectancy before costs.",
        changes_authorized=(
            "Restrict levels to PDH/PDL, require 12x relative volume, and use 1.25R target.",
        ),
        changes_rejected=("Changes to frozen allocator, costs, and OOS policy.",),
        next_action="Run v1.1.0 on the same frozen development interval.",
        recorded_at=datetime.fromisoformat("2026-08-14T21:02:11+05:30"),
        git_commit="ed1c40f",
    ),
    ResearchDecisionRecord(
        decision_id="strategy1-v1.1-development-review",
        source_run_ids=(
            "strategy1-development-2016-12-29-to-2017-03-29-1.1.0-6d8908a13d44",
        ),
        strategy_id=RESEARCH_SCOPE_ID,
        strategy_version="1.1.0",
        research_scope_id=RESEARCH_SCOPE_ID,
        decision="ADVANCE_UNCHANGED_TO_OOS_001",
        diagnosis="Promising but small-sample development evidence required untouched data.",
        changes_authorized=(),
        changes_rejected=("Further threshold tuning on the development sample.",),
        next_action="Run v1.1.0 unchanged on the governed OOS-001 window.",
        recorded_at=datetime.fromisoformat("2026-08-14T22:02:36+05:30"),
        git_commit="e5a1ab8",
    ),
    ResearchDecisionRecord(
        decision_id="strategy1-v1.1-oos-001-review",
        source_run_ids=(
            "strategy1-oos-2017-03-29-to-2017-06-29-1.1.0-373f87db4491",
        ),
        strategy_id=RESEARCH_SCOPE_ID,
        strategy_version="1.1.0",
        research_scope_id=RESEARCH_SCOPE_ID,
        decision="OOS_001_FAILED_DO_NOT_RETEST",
        diagnosis="OOS-001 failed materially and is now inspected research evidence.",
        changes_authorized=("Repair platform integration before further strategy research.",),
        changes_rejected=("Reuse OOS-001 as a fresh test or tune directly on it.",),
        next_action="Complete integration repair before any Strategy 1 v1.2 research.",
        recorded_at=datetime.fromisoformat("2026-08-14T22:33:22+05:30"),
        git_commit="78f74d0",
    ),
)


class _UnusedMarginProvider:
    """Valid protocol implementation used only when a window generates zero requests."""

    def quote(
        self,
        candidate: AllocationCandidate,
        state: PortfolioState,
    ) -> MarginRequirementQuote:
        raise RuntimeError("margin provider must not be called when no requests exist")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Strategy-1 research runner. Default mode runs/re-runs the frozen development window. "
            "Use next-oos only after reviewing development or after authorizing the prior OOS."
        )
    )
    parser.add_argument(
        "--mode",
        choices=("development", "next-oos", "authorize-oos"),
        default="development",
    )
    parser.add_argument(
        "--brokerage-plan",
        choices=(BrokeragePlan.PLUS.value, BrokeragePlan.PRO_PLUS.value),
        default=DEFAULT_BROKERAGE_PLAN.value,
    )
    parser.add_argument(
        "--slippage-bps",
        default=str(DEFAULT_SLIPPAGE_BPS),
        help="Adverse slippage applied independently to each simulated fill.",
    )
    parser.add_argument(
        "--window-id",
        default=None,
        help="TESTED OOS window to consume+authorize. Optional in authorize-oos mode.",
    )
    return parser.parse_args()


def git_output(*args: str) -> str:
    return subprocess.check_output(
        ["git", *args],
        cwd=ROOT,
        text=True,
    ).strip()


def _substantive_unstaged_diff(path: str) -> bool:
    result = subprocess.run(
        ["git", "diff", "--quiet", "--ignore-space-at-eol", "--", path],
        cwd=ROOT,
        check=False,
    )
    return result.returncode != 0


def repository_reproducibility_state() -> tuple[str, str, tuple[str, ...], str]:
    """Allow the runner itself to be untracked/modified, but reject other real source dirt.

    The Git commit plus exact runner SHA-256 is persisted in every run manifest. Line-ending-only
    tracked changes are tolerated and explicitly reported because they do not change Python
    semantics. Any staged change or any other substantive unstaged/untracked file blocks research.
    """

    git_commit = git_output("rev-parse", "HEAD")
    runner_sha256 = hashlib.sha256(RUNNER_PATH.read_bytes()).hexdigest()
    status_lines = git_output("status", "--porcelain").splitlines()

    blocking: list[str] = []
    eol_only: list[str] = []
    runner_state = "TRACKED_CLEAN"

    for line in status_lines:
        if not line:
            continue
        status = line[:2]
        path = line[3:].strip()
        if " -> " in path:
            path = path.split(" -> ", 1)[1]

        if path == RUNNER_RELATIVE_PATH:
            runner_state = (
                "UNTRACKED"
                if status == "??"
                else "MODIFIED"
                if "M" in status
                else status.strip() or "CHANGED"
            )
            # Staging the runner is still safe, but any staged non-runner file is not.
            continue

        if status == "??":
            blocking.append(line)
            continue

        # Anything staged outside the runner is a different repository state than HEAD.
        if status[0] != " ":
            blocking.append(line)
            continue

        # Unstaged tracked changes that are only EOL normalization are tolerated.
        if status[1] != " ":
            if _substantive_unstaged_diff(path):
                blocking.append(line)
            else:
                eol_only.append(path)

    if blocking:
        raise RuntimeError(
            "Reproducible research blocked by substantive repository changes outside the runner:\n"
            + "\n".join(blocking)
            + "\nCommit/revert those changes before running a genuine backtest."
        )

    return git_commit, runner_sha256, tuple(sorted(eol_only)), runner_state


def as_ist_midnight(value: date) -> datetime:
    return datetime.combine(value, time.min, tzinfo=IST)


def market_date(value: datetime) -> date:
    if value.tzinfo is None or value.utcoffset() is None:
        raise ValueError("market timestamp must be timezone-aware")
    return value.astimezone(IST).date()


def decimal_text(value: Decimal | None) -> str:
    return "" if value is None else format(value, "f")


def profit_factor_text(report: ReportBundle) -> str:
    value = report.performance.net_profit_factor
    if value.is_unbounded:
        return "INF"
    if value.is_undefined:
        return "UNDEFINED"
    return decimal_text(value.value)


def finite_profit_factor(report: ReportBundle) -> float | None:
    value = report.performance.net_profit_factor
    return float(value.value) if value.value is not None else None


def ceil_notional_bucket(price: Decimal) -> int | None:
    if not price.is_finite() or price <= 0:
        raise ValueError("decision price must be finite and positive")
    required = max(Decimal(MIN_NOTIONAL), price)
    bucket = int(
        (required / Decimal(NOTIONAL_INCREMENT)).to_integral_value(
            rounding=ROUND_CEILING
        )
    ) * NOTIONAL_INCREMENT
    if bucket > MAX_NOTIONAL:
        return None
    return bucket


def decision_quantity(notional: int, price: Decimal) -> int:
    quantity = int(Decimal(notional) // price)
    return max(1, quantity)


def ready_bar_for_symbol(
    connection: duckdb.DuckDBPyConnection,
    parquet_path: Path,
    warmup_bars: int,
) -> datetime | None:
    row = connection.execute(
        """
        SELECT date::VARCHAR
        FROM read_parquet(?)
        ORDER BY date
        LIMIT 1 OFFSET ?
        """,
        [str(parquet_path), warmup_bars - 1],
    ).fetchone()
    if row is None or row[0] is None:
        return None
    parsed = datetime.fromisoformat(row[0])
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        parsed = parsed.replace(tzinfo=UTC)
    return parsed.astimezone(IST)


def derive_earliest_strategy_ready(
    *,
    coverages: tuple[SymbolCoverage, ...],
    strategy: LiquidityShockReclaimStrategy,
) -> tuple[str, datetime]:
    excluded = set(STANDARD_OOS_PARTITION_POLICY.excluded_reference_symbols)
    candidates = [
        coverage
        for coverage in coverages
        if coverage.symbol not in excluded
        and coverage.row_count >= strategy.warmup_bars
        and coverage.first_timestamp is not None
        and coverage.last_timestamp is not None
    ]
    if not candidates:
        raise RuntimeError("No equity has enough rows for Strategy-1 warm-up.")

    best: tuple[str, datetime] | None = None
    with duckdb.connect() as connection:
        connection.execute("SET TimeZone = 'Asia/Kolkata'")
        for index, coverage in enumerate(candidates, start=1):
            ready_bar = ready_bar_for_symbol(
                connection,
                DATA_PATH / f"{coverage.symbol}.parquet",
                strategy.warmup_bars,
            )
            if ready_bar is None:
                continue
            available_at = bar_available_at(ready_bar)
            candidate = (coverage.symbol, available_at)
            if best is None or (candidate[1], candidate[0]) < (best[1], best[0]):
                best = candidate
            if index % 100 == 0:
                print(f"Warm-up readiness scan: {index}/{len(candidates)} symbols")

    if best is None:
        raise RuntimeError("Could not derive an earliest Strategy-1-ready timestamp.")
    return best


def run_real_causality_gate(
    *,
    store: ParquetMarketDataStore,
    coverages: tuple[SymbolCoverage, ...],
    strategy: LiquidityShockReclaimStrategy,
    allowed_end_exclusive: date,
) -> StrategyCausalityReport:
    """Prove Strategy 1 on a real signal drawn only from allowed research history."""
    allowed_end = as_ist_midnight(allowed_end_exclusive)
    excluded = set(STANDARD_OOS_PARTITION_POLICY.excluded_reference_symbols)
    best: tuple[datetime, str, Signal, pl.DataFrame] | None = None
    for coverage in sorted(coverages, key=lambda item: item.symbol):
        if (
            coverage.symbol in excluded
            or coverage.first_timestamp is None
            or coverage.row_count < strategy.warmup_bars + 3
        ):
            continue
        candles = store.load_candles(
            coverage.symbol,
            coverage.first_timestamp,
            allowed_end,
        )
        if candles.height < strategy.warmup_bars + 3:
            continue
        signals = [
            signal
            for signal in strategy.generate_signals(candles)
            if signal.timestamp < allowed_end
        ]
        if signals:
            first = signals[0]
            candidate = (first.timestamp, coverage.symbol, first, candles)
            if best is None or candidate[:2] < best[:2]:
                best = candidate

    if best is None:
        raise RuntimeError(
            "Causality preflight found no real Strategy-1 signal in development or "
            "TRAINING_ALLOWED history."
        )
    _, symbol, target_signal, source = best
    availability = [bar_available_at(value) for value in source["timestamp"].to_list()]
    knowledge_rows = sum(value <= target_signal.timestamp for value in availability)
    future_rows = min(source.height, knowledge_rows + 8)
    start = max(0, knowledge_rows - strategy.warmup_bars)
    bounded = source.slice(start, future_rows - start)

    while target_signal not in strategy.generate_signals(bounded) and start > 0:
        start = max(0, start - 250)
        bounded = source.slice(start, future_rows - start)

    bounded_signals = strategy.generate_signals(bounded)
    if target_signal not in bounded_signals:
        raise RuntimeError("Causality preflight could not preserve the located signal identity.")
    report = assert_strategy_prefix_invariant(strategy, bounded)
    if report.full_signal_count < 1 or report.tested_prefix_count <= 2:
        raise RuntimeError(
            "Causality preflight is vacuous: it requires a real signal and more than two "
            "semantic prefixes."
        )
    if report.symbol != symbol:
        raise RuntimeError("Causality preflight reported an unexpected symbol.")
    return report


def build_request(signal: Signal, score: MLScore) -> BacktestTradeRequest | None:
    feature = signal.feature_snapshot
    decision_price = Decimal(str(feature["confirmation_close"]))
    if ceil_notional_bucket(decision_price) is None:
        return None
    target_notional = score.recommended_notional

    candidate = AllocationCandidate(
        order_intent=OrderIntent(
            signal=signal,
            timestamp=signal.timestamp,
            quantity=decision_quantity(target_notional, decision_price),
            requested_notional=target_notional,
            order_type=OrderType.MARKET,
        ),
        ml_score=score,
    )

    return BacktestTradeRequest(
        candidate=candidate,
        dynamic_exit_policy=DynamicExitPolicySpec(
            policy_id=DYNAMIC_EXIT_POLICY_ID,
            parameters={
                "initial_stop_price": feature["stop_reference_price"],
                "hard_target_r": feature["trailing_hard_target_r"],
                "breakeven_trigger_r": feature["trailing_breakeven_trigger_r"],
                "breakeven_stop_r": feature["trailing_breakeven_stop_r"],
                "profit_lock_trigger_r": feature["trailing_profit_lock_trigger_r"],
                "profit_lock_stop_r": feature["trailing_profit_lock_stop_r"],
                "trailing_distance_r": feature["trailing_distance_r"],
                "maximum_hold_minutes": feature["maximum_hold_minutes"],
                "latest_exit_time": time.fromisoformat(
                    str(signal.strategy_parameters["latest_exit_time"])
                ),
            },
        ),
    )


def generate_requests(
    *,
    store: ParquetMarketDataStore,
    coverages: tuple[SymbolCoverage, ...],
    strategy: LiquidityShockReclaimStrategy,
    scorer: TradeScorer,
    window_start: date,
    window_end: date,
) -> tuple[
    list[BacktestTradeRequest],
    tuple[str, ...],
    tuple[date, ...],
    int,
    int,
]:
    """Scan every historically available equity; strategy rules decide signal eligibility."""

    scanned_symbols = select_historically_available_equities(
        coverages,
        start_date=window_start,
        end_date=window_end,
    )
    coverage_by_symbol = {coverage.symbol: coverage for coverage in coverages}
    start_dt = as_ist_midnight(window_start)
    end_dt = as_ist_midnight(window_end)

    requests: list[BacktestTradeRequest] = []
    trading_dates: set[date] = set()
    skipped_expensive = 0
    raw_signal_count = 0

    print(f"\nFull historical scan universe: {len(scanned_symbols)} equities")
    print(f"Scored window: {window_start} -> {window_end}")

    for index, symbol in enumerate(scanned_symbols, start=1):
        coverage = coverage_by_symbol[symbol]
        if coverage.first_timestamp is None:
            continue

        # Loading from each symbol's own first row preserves causal warm-up for later listings.
        candles = store.load_candles(
            symbol,
            coverage.first_timestamp,
            end_dt,
        )
        if candles.is_empty():
            continue

        in_window = candles.filter(
            (pl.col("timestamp") >= start_dt) & (pl.col("timestamp") < end_dt)
        )
        if not in_window.is_empty():
            trading_dates.update(
                in_window.select(pl.col("timestamp").dt.date().unique()).to_series().to_list()
            )

        signals = [
            signal
            for signal in strategy.generate_signals(candles)
            if start_dt <= signal.timestamp < end_dt
        ]
        raw_signal_count += len(signals)

        built = score_and_build_requests(signals, scorer, build_request)
        accepted = len(built)
        skipped_expensive += len(signals) - accepted
        requests.extend(built)

        if signals:
            print(
                f"  {symbol}: signals={len(signals)} requests={accepted} "
                f"rows_loaded={candles.height:,}"
            )
        elif index % 25 == 0 or index == len(scanned_symbols):
            print(
                f"  progress {index}/{len(scanned_symbols)} | "
                f"signals={raw_signal_count} | requests={len(requests)}"
            )

    requests.sort(
        key=lambda request: (
            request.candidate.order_intent.timestamp,
            request.candidate.identity,
        )
    )
    return (
        requests,
        tuple(scanned_symbols),
        tuple(sorted(trading_dates)),
        skipped_expensive,
        raw_signal_count,
    )


def required_margin_pairs(
    requests: Iterable[BacktestTradeRequest],
) -> set[tuple[str, Side]]:
    return {
        (
            request.candidate.order_intent.signal.symbol,
            request.candidate.order_intent.signal.side,
        )
        for request in requests
    }


def latest_margin_snapshot() -> HistoricalMarginSnapshot | None:
    if not MARGIN_SNAPSHOT_DIR.exists():
        return None
    paths = sorted(MARGIN_SNAPSHOT_DIR.glob("*.json"), reverse=True)
    if paths:
        try:
            return load_historical_margin_snapshot(paths[0])
        except Exception as error:
            raise RuntimeError(
                f"Latest margin snapshot is corrupt or unreadable: {paths[0]}"
            ) from error
    return None


def snapshot_entries_by_pair(
    snapshot: HistoricalMarginSnapshot | None,
) -> dict[tuple[str, Side], object]:
    if snapshot is None:
        return {}
    return {(entry.symbol, entry.side): entry for entry in snapshot.entries}


def capture_or_expand_margin_snapshot(
    *,
    requests: list[BacktestTradeRequest],
    captured_at: datetime,
) -> HistoricalMarginSnapshot | None:
    pairs = required_margin_pairs(requests)
    if not pairs:
        return None

    previous = latest_margin_snapshot()
    requested_as_of = captured_at.astimezone(IST).date()
    if previous is not None and previous.source_as_of_date != requested_as_of:
        print(
            "Existing margin snapshot is stale for the requested fixed-current policy: "
            f"snapshot_as_of={previous.source_as_of_date} requested_as_of={requested_as_of}. "
            "A fresh complete snapshot is required."
        )
        previous = None
    entries = snapshot_entries_by_pair(previous)
    missing = sorted(pairs - set(entries), key=lambda item: (item[0], item[1].value))
    if not missing and previous is not None:
        print(f"Using existing broker-derived margin snapshot: {previous.snapshot_id}")
        return previous

    print(
        f"\nMargin evidence: {len(pairs)} required symbol/side pairs; "
        f"{len(missing)} new broker quotes required."
    )

    credentials = load_smartapi_credentials()
    instrument_master = fetch_instrument_master()
    if HISTORICAL_SYMBOL_ALIAS_PATH.exists():
        aliases = json.loads(HISTORICAL_SYMBOL_ALIAS_PATH.read_text(encoding="utf-8"))
        if not isinstance(aliases, Mapping):
            raise RuntimeError("historical symbol alias file must contain a JSON object")
        instrument_master = instrument_master.with_aliases(aliases)
        print(
            "Loaded explicit historical symbol aliases: "
            f"{len(aliases)} from {HISTORICAL_SYMBOL_ALIAS_PATH}"
        )
    broker = AngelOneBroker(instrument_master)
    session = None
    unresolved: list[str] = []

    try:
        session = broker.authenticate(credentials, captured_at)

        live_margin = AngelOneLiveMarginProvider(
            broker.authenticated_client(),
            instrument_master,
        )

        for index, (symbol, side) in enumerate(missing, start=1):
            try:
                instrument = instrument_master.resolve(symbol)
                quote = broker.get_ltp(symbol, captured_at)
                ltp = quote.ltp
                lot_size = instrument.lot_size

                raw_quantity = max(
                    lot_size,
                    int(Decimal(MARGIN_REFERENCE_NOTIONAL) // ltp),
                )
                quantity = max(lot_size, (raw_quantity // lot_size) * lot_size)
                reference_notional = ltp * quantity

                calibration_signal = Signal(
                    strategy_id="historical-margin-calibration",
                    strategy_version="1",
                    symbol=symbol,
                    timestamp=captured_at,
                    side=side,
                )
                calibration_score = MLScore(
                    model_version="historical-margin-calibration-v1",
                    quality_score=0.5,
                    calibrated_probability=0.5,
                    predicted_net_return=0.0,
                    recommended_notional=MARGIN_REFERENCE_NOTIONAL,
                )
                calibration_candidate = AllocationCandidate(
                    order_intent=OrderIntent(
                        signal=calibration_signal,
                        timestamp=captured_at,
                        quantity=quantity,
                        requested_notional=MARGIN_REFERENCE_NOTIONAL,
                        order_type=OrderType.MARKET,
                    ),
                    ml_score=calibration_score,
                )
                margin_quote = live_margin.quote(
                    calibration_candidate,
                    PortfolioState(capital_limit=INITIAL_CAPITAL),
                )
                entry = create_margin_snapshot_entry(
                    symbol,
                    side,
                    reference_notional,
                    margin_quote.required_margin,
                )
                entries[(symbol, side)] = entry
                print(
                    f"  margin {index}/{len(missing)} {symbol}/{side.value}: "
                    f"required≈₹{entry.broker_required_margin:,.2f} per "
                    f"₹{entry.reference_notional:,.2f} exposure | "
                    f"margin={entry.required_margin_fraction * 100:.2f}% | "
                    f"implied_exposure≈{Decimal('1') / entry.required_margin_fraction:.2f}x"
                )
            except BrokerSystemicError as error:
                raise RuntimeError(
                    "Broker-derived margin capture aborted after the first representative "
                    "shared/systemic failure; no fallback margin was used: "
                    f"{symbol}/{side.value}: {error}"
                ) from error
            except Exception as error:
                unresolved.append(
                    f"{symbol}/{side.value}: {type(error).__name__}: {error}"
                )
    finally:
        if session is not None:
            broker.logout(session)

    if unresolved:
        raise RuntimeError(
            "Full-universe scan produced signal(s) whose broker-derived margin could not be "
            "resolved. They are NOT silently discarded:\n" + "\n".join(unresolved)
        )

    snapshot = create_historical_margin_snapshot(
        snapshot_id=f"strategy1-margin-{captured_at:%Y%m%d-%H%M%S}",
        captured_at=captured_at,
        source_as_of_date=captured_at.date(),
        entries=tuple(entries.values()),
    )
    MARGIN_SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    output = MARGIN_SNAPSHOT_DIR / f"{snapshot.snapshot_id}.json"
    save_historical_margin_snapshot(snapshot, output)
    print(f"Saved expanded broker-derived margin snapshot: {output}")
    return snapshot


def get_or_create_standard_plan(
    *,
    registry: OOSRegistry,
    data_start: date,
    data_end: date,
    earliest_oos_start: date,
    git_commit: str,
    occurred_at: datetime,
) -> OOSPlan:
    try:
        plan = registry.get_plan(RESEARCH_SCOPE_ID, PLAN_ID)
    except LookupError:
        proposed = create_standard_oos_plan(
            research_scope_id=RESEARCH_SCOPE_ID,
            plan_id=PLAN_ID,
            strategy_ids=(RESEARCH_SCOPE_ID,),
            data_start_date=data_start,
            data_end_exclusive=data_end,
            earliest_oos_start_date=earliest_oos_start,
            audit_context=OOSAuditContext(
                event_id=f"strategy1-plan-{uuid4().hex}",
                occurred_at=occurred_at,
                git_commit=git_commit,
            ),
            policy=STANDARD_OOS_PARTITION_POLICY,
        )
        plan = registry.create_plan(proposed)
        print(
            f"\nOOS PLAN FROZEN: {len(plan.oos_windows)} ordinary windows | "
            f"sealed={plan.sealed_holdout.start_date}->{plan.sealed_holdout.end_date}"
        )
        return plan

    expected = create_standard_oos_plan(
        research_scope_id=RESEARCH_SCOPE_ID,
        plan_id=PLAN_ID,
        strategy_ids=(RESEARCH_SCOPE_ID,),
        data_start_date=data_start,
        data_end_exclusive=data_end,
        earliest_oos_start_date=earliest_oos_start,
        audit_context=plan.creation_audit,
        policy=STANDARD_OOS_PARTITION_POLICY,
    )
    immutable_fields = (
        "protocol_version",
        "partition_policy",
        "data_start_date",
        "data_end_exclusive",
        "development_start_date",
        "development_end_exclusive",
        "sealed_holdout_start_date",
        "sealed_holdout_end_exclusive",
    )
    for field in immutable_fields:
        if getattr(plan, field) != getattr(expected, field):
            raise RuntimeError(
                f"Persisted OOS plan does not match frozen standard research policy: {field}"
            )
    if tuple((w.window_id, w.start_date, w.end_date) for w in plan.oos_windows) != tuple(
        (w.window_id, w.start_date, w.end_date) for w in expected.oos_windows
    ):
        raise RuntimeError("Persisted OOS plan window boundaries do not match frozen policy.")
    return plan


def development_dates(
    *,
    data_start: date,
    data_end: date,
    ready_at: datetime,
    registry: OOSRegistry,
) -> tuple[date, date, date, OOSPlan | None]:
    evaluation_start = market_date(ready_at)
    proposed_oos_start = shift_calendar_months(
        evaluation_start,
        DEVELOPMENT_EVALUATION_MONTHS,
    )
    sealed_start = shift_calendar_months(
        data_end,
        -STANDARD_OOS_PARTITION_POLICY.sealed_holdout_months,
    )
    if proposed_oos_start >= sealed_start:
        raise RuntimeError("Insufficient pre-holdout history for development plus ordinary OOS.")

    try:
        plan = registry.get_plan(RESEARCH_SCOPE_ID, PLAN_ID)
    except LookupError:
        plan = None
        evaluation_end = proposed_oos_start
    else:
        if plan.data_start_date != data_start or plan.data_end_exclusive != data_end:
            raise RuntimeError("Current equity data horizon differs from the frozen OOS plan.")
        evaluation_end = plan.development_end_exclusive
        if evaluation_start >= evaluation_end:
            raise RuntimeError(
                "Current Strategy-1 warm-up leaves no scored development interval inside the "
                "already-frozen development range."
            )

    return evaluation_start, evaluation_end, proposed_oos_start, plan


def create_backtester(
    *,
    store: ParquetMarketDataStore,
    margin_snapshot: HistoricalMarginSnapshot | None,
    slippage_bps: Decimal,
) -> HistoricalBacktester:
    margin_provider = (
        HistoricalMarginRequirementProvider(margin_snapshot)
        if margin_snapshot is not None
        else _UnusedMarginProvider()
    )
    return HistoricalBacktester(
        store,
        margin_provider,
        execution_simulator=HistoricalExecutionSimulator(
            slippage_model=FixedBasisPointsSlippage(slippage_bps)
        ),
        exit_policy_resolvers=(RMultipleTrailingExitPolicyResolver(),),
    )


def stable_run_id(
    *,
    phase: str,
    strategy_version: str,
    window_start: date,
    window_end: date,
    git_commit: str,
    runner_sha256: str,
    brokerage_plan: BrokeragePlan,
    slippage_bps: Decimal,
    run_input_fingerprint: str,
) -> str:
    material = "|".join(
        (
            phase,
            strategy_version,
            window_start.isoformat(),
            window_end.isoformat(),
            git_commit,
            runner_sha256,
            brokerage_plan.value,
            str(slippage_bps),
            run_input_fingerprint,
        )
    )
    suffix = hashlib.sha256(material.encode("utf-8")).hexdigest()[:12]
    return (
        f"strategy1-{phase}-{window_start.isoformat()}-to-{window_end.isoformat()}-"
        f"{strategy_version}-{suffix}"
    )


def write_json_model(path: Path, model: object) -> None:
    if path.exists():
        raise FileExistsError(f"artifact already exists: {path}")
    if hasattr(model, "model_dump_json"):
        text = model.model_dump_json(indent=2, exclude_computed_fields=True)
    else:
        raise TypeError("model must support model_dump_json")
    path.write_text(text, encoding="utf-8", newline="\n")


def write_run_manifest(
    *,
    path: Path,
    phase: str,
    report: ReportBundle,
    git_commit: str,
    runner_sha256: str,
    runner_git_state: str,
    eol_only_paths: tuple[str, ...],
    data_start: date,
    data_end: date,
    ready_symbol: str,
    ready_at: datetime,
    scanned_symbols: tuple[str, ...],
    raw_signal_count: int,
    request_count: int,
    trading_day_count: int,
    skipped_expensive: int,
    slippage_bps: Decimal,
    margin_snapshot_id: str | None,
    causality: StrategyCausalityReport,
    reproducibility: dict[str, object],
) -> None:
    payload = {
        "phase": phase,
        "completed_at": report.provenance.generated_at.isoformat(),
        "research_scope_id": RESEARCH_SCOPE_ID,
        "plan_id": report.provenance.plan_id,
        "window_id": report.provenance.window_id,
        "strategy_id": RESEARCH_SCOPE_ID,
        "strategy_version": dict(report.provenance.strategy_versions).get(
            RESEARCH_SCOPE_ID, "UNKNOWN"
        ),
        "git_commit": git_commit,
        "runner_sha256": runner_sha256,
        "runner_git_state": runner_git_state,
        "line_ending_only_tracked_paths": list(eol_only_paths),
        "data_horizon": {
            "start": data_start.isoformat(),
            "end_exclusive": data_end.isoformat(),
        },
        "earliest_strategy_ready": {
            "symbol": ready_symbol,
            "available_at": ready_at.isoformat(),
        },
        "scored_window": {
            "start": report.provenance.window_start.date().isoformat(),
            "end_exclusive": report.provenance.window_end.date().isoformat(),
        },
        "partition_policy": STANDARD_OOS_PARTITION_POLICY.model_dump(mode="json"),
        "universe_policy_id": STANDARD_OOS_PARTITION_POLICY.universe_policy_id,
        "scanned_symbol_count": len(scanned_symbols),
        "scanned_symbols": list(scanned_symbols),
        "raw_signal_count": raw_signal_count,
        "request_count": request_count,
        "trading_day_count": trading_day_count,
        "skipped_price_above_max_notional": skipped_expensive,
        "bootstrap_model_version": BOOTSTRAP_MODEL_VERSION,
        "initial_capital": str(INITIAL_CAPITAL),
        "brokerage_plan": report.provenance.brokerage_plan,
        "slippage_bps_per_fill": str(slippage_bps),
        "margin_snapshot_id": margin_snapshot_id,
        "dynamic_exit_policy_id": DYNAMIC_EXIT_POLICY_ID,
        "run_input_fingerprint": reproducibility["run_input_fingerprint"],
        "reproducibility": reproducibility,
        "causality_gate": {
            "gate_version": causality.gate_version,
            "strategy_id": causality.strategy_id,
            "strategy_version": causality.strategy_version,
            "symbol": causality.symbol,
            "tested_prefix_count": causality.tested_prefix_count,
            "full_signal_count": causality.full_signal_count,
            "row_count": causality.row_count,
            "pass": True,
        },
        "acceptance": report.acceptance.model_dump(mode="json"),
    }
    path.write_text(
        json.dumps(payload, indent=2, sort_keys=True),
        encoding="utf-8",
        newline="\n",
    )


def _run_artifacts() -> list[tuple[ReportBundle, dict[str, object], Path]]:
    discovery = discover_research_runs(RESULTS_ROOT)
    return [
        (run.report, dict(run.manifest), run.directory)
        for run in discovery.completed
    ]


def _excel_value(value: object) -> object:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.astimezone(IST).replace(tzinfo=None)
    if isinstance(value, date):
        return value
    if isinstance(value, (tuple, list, dict)):
        return json.dumps(value, default=str, separators=(",", ":"))
    return value


def _write_rows(sheet, rows: list[dict[str, object]]) -> None:
    if not rows:
        sheet.append(["No records"])
        return
    headers = list(rows[0])
    sheet.append(headers)
    for row in rows:
        sheet.append([_excel_value(row.get(header)) for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for column_index, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row_index in range(2, min(sheet.max_row, 500) + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 10), 34
        )


def _table_rows(
    report: ReportBundle,
    table_name: str,
    *,
    attempt: int,
    phase: str,
    window_id: str | None,
) -> list[dict[str, object]]:
    frame = report_tables(report)[table_name]
    rows = []
    for row in frame.to_dicts():
        rows.append(
            {
                "attempt": attempt,
                "phase": phase,
                "window_id": window_id,
                "run_id": report.provenance.run_id,
                **row,
            }
        )
    return rows


def _run_history_rows(
    artifacts: list[tuple[ReportBundle, dict[str, object], Path]],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for attempt, (report, manifest, _) in enumerate(artifacts, start=1):
        p = report.performance
        rows.append(
            {
                "attempt": attempt,
                "phase": manifest.get("phase"),
                "window_id": report.provenance.window_id,
                "strategy_version": manifest.get("strategy_version"),
                "generated_at": report.provenance.generated_at,
                "run_id": report.provenance.run_id,
                "git_commit": report.provenance.git_commit,
                "runner_sha256": manifest.get("runner_sha256"),
                "window_start": report.provenance.window_start,
                "window_end": report.provenance.window_end,
                "scanned_symbols": manifest.get("scanned_symbol_count"),
                "raw_signals": manifest.get("raw_signal_count"),
                "requests": manifest.get("request_count"),
                "actual_trades": p.actual_trade_count,
                "shadow_trades": report.shadow_metrics.shadow_trade_count,
                "ending_capital": p.ending_capital,
                "net_profit": p.net_profit,
                "total_return": p.total_return,
                "cagr": p.cagr,
                "win_rate": p.win_rate,
                "profit_factor": finite_profit_factor(report),
                "profit_factor_display": profit_factor_text(report),
                "average_net_return": p.average_net_return_per_trade,
                "max_drawdown_pct": p.maximum_realized_drawdown_pct,
                "trades_per_day": p.actual_trades_per_day,
                "total_costs": p.total_costs,
                "cagr_pass": report.acceptance.cagr_pass,
                "win_rate_pass": report.acceptance.win_rate_pass,
                "profit_factor_pass": report.acceptance.profit_factor_pass,
                "average_net_return_pass": report.acceptance.average_net_return_pass,
                "frequency_target_met": report.acceptance.frequency_target_met,
                "hard_quantitative_targets_pass": (
                    report.acceptance.hard_quantitative_targets_pass
                ),
            }
        )
    return rows



def _research_decision_rows() -> list[dict[str, object]]:
    return [row.model_dump(mode="json") for row in RESEARCH_DECISION_HISTORY]


def rebuild_master_json(plan: OOSPlan | None) -> None:
    """Write the single cumulative machine-readable upload artifact atomically."""
    artifacts = _run_artifacts()
    if not artifacts:
        return

    run_payloads: list[dict[str, object]] = []
    for attempt, (report, manifest, run_directory) in enumerate(artifacts, start=1):
        report_path = run_directory / "report_bundle.json"
        manifest_path = run_directory / "run_manifest.json"
        run_payloads.append(
            {
                "attempt": attempt,
                "phase": manifest.get("phase"),
                "window_id": report.provenance.window_id,
                "run_id": report.provenance.run_id,
                "strategy_version": manifest.get("strategy_version"),
                "run_directory": run_directory.relative_to(RESULTS_ROOT).as_posix(),
                "report_sha256": hashlib.sha256(report_path.read_bytes()).hexdigest(),
                "manifest_sha256": hashlib.sha256(manifest_path.read_bytes()).hexdigest(),
                "manifest": manifest,
                "report": report.model_dump(mode="json"),
            }
        )

    payload = {
        "schema_version": "1",
        "research_scope_id": RESEARCH_SCOPE_ID,
        "generated_at": datetime.now(IST).replace(microsecond=0).isoformat(),
        "upload_contract": {
            "machine_readable": MASTER_JSON_PATH.name,
            "human_analysis": MASTER_EXCEL_PATH.name,
            "instruction": (
                "For review, upload only this cumulative JSON, the cumulative Excel workbook, "
                "and the terminal output. Per-run Parquet/JSON/PNG artifacts remain internal "
                "reproducibility evidence and do not need to be uploaded."
            ),
        },
        "research_decision_history": _research_decision_rows(),
        "run_history": _run_history_rows(artifacts),
        "gate_history": _gate_history_rows(artifacts),
        "oos_plan": _oos_plan_rows(plan),
        "runs": run_payloads,
    }

    MASTER_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    temp_path = MASTER_JSON_PATH.with_suffix(".tmp.json")
    temp_path.write_text(
        json.dumps(payload, indent=2, sort_keys=True, default=str),
        encoding="utf-8",
        newline="\n",
    )
    os.replace(temp_path, MASTER_JSON_PATH)


def _gate_history_rows(
    artifacts: list[tuple[ReportBundle, dict[str, object], Path]],
) -> list[dict[str, object]]:
    rows = []
    for attempt, (report, manifest, _) in enumerate(artifacts, start=1):
        gate = manifest.get("causality_gate", {})
        rows.append(
            {
                "attempt": attempt,
                "phase": manifest.get("phase"),
                "window_id": report.provenance.window_id,
                "run_id": report.provenance.run_id,
                "causality_gate_version": (
                    gate.get("gate_version") if isinstance(gate, dict) else None
                ),
                "causality_pass": gate.get("pass") if isinstance(gate, dict) else None,
                "full_universe_policy": (
                    manifest.get("universe_policy_id")
                    == STANDARD_OOS_PARTITION_POLICY.universe_policy_id
                ),
                "scanned_symbol_count": manifest.get("scanned_symbol_count"),
                "references_excluded": all(
                    symbol not in set(manifest.get("scanned_symbols", []))
                    for symbol in STANDARD_OOS_PARTITION_POLICY.excluded_reference_symbols
                ),
                "cagr_gt_20pct": report.acceptance.cagr_pass,
                "win_rate_gt_50pct": report.acceptance.win_rate_pass,
                "net_pf_gt_2": report.acceptance.profit_factor_pass,
                "avg_net_return_gt_0_5pct": report.acceptance.average_net_return_pass,
                "portfolio_frequency_ge_2_per_day": report.acceptance.frequency_target_met,
                "hard_quantitative_targets_pass": (
                    report.acceptance.hard_quantitative_targets_pass
                ),
            }
        )
    return rows


def _oos_plan_rows(plan: OOSPlan | None) -> list[dict[str, object]]:
    if plan is None:
        return [{"status": "OOS plan not created yet"}]
    rows = [
        {
            "window_id": "development",
            "start_date": plan.development_start_date,
            "end_date": plan.development_end_exclusive,
            "state": "DEVELOPMENT",
            "months_target": None,
            "policy_id": plan.partition_policy.policy_id if plan.partition_policy else None,
        }
    ]
    rows.extend(
        {
            "window_id": window.window_id,
            "start_date": window.start_date,
            "end_date": window.end_date,
            "state": window.state.value,
            "months_target": STANDARD_OOS_PARTITION_POLICY.target_window_months,
            "policy_id": (
                plan.partition_policy.policy_id if plan.partition_policy else None
            ),
        }
        for window in plan.oos_windows
    )
    rows.append(
        {
            "window_id": plan.sealed_holdout.window_id,
            "start_date": plan.sealed_holdout.start_date,
            "end_date": plan.sealed_holdout.end_date,
            "state": plan.sealed_holdout.state.value,
            "months_target": STANDARD_OOS_PARTITION_POLICY.sealed_holdout_months,
            "policy_id": plan.partition_policy.policy_id if plan.partition_policy else None,
        }
    )
    return rows


def _add_dashboard_chart(
    dashboard,
    history,
    *,
    title: str,
    column_header: str,
    anchor: str,
    chart_type: str = "line",
) -> None:
    if history.max_row <= 2:
        return
    headers = {cell.value: cell.column for cell in history[1]}
    column = headers.get(column_header)
    attempt_column = headers.get("attempt")
    if column is None or attempt_column is None:
        return

    chart = LineChart() if chart_type == "line" else BarChart()
    chart.title = title
    chart.height = 7
    chart.width = 12
    chart.x_axis.title = "Research attempt"
    chart.y_axis.title = column_header

    data = Reference(history, min_col=column, min_row=1, max_row=history.max_row)
    categories = Reference(
        history, min_col=attempt_column, min_row=2, max_row=history.max_row
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    dashboard.add_chart(chart, anchor)


def rebuild_master_workbook(plan: OOSPlan | None) -> None:
    artifacts = _run_artifacts()
    if not artifacts:
        return

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"

    run_rows = _run_history_rows(artifacts)
    history = workbook.create_sheet("Run History")
    _write_rows(history, run_rows)

    gates = workbook.create_sheet("Gate History")
    _write_rows(gates, _gate_history_rows(artifacts))

    oos_sheet = workbook.create_sheet("OOS Plan")
    _write_rows(oos_sheet, _oos_plan_rows(plan))

    decisions = workbook.create_sheet("Research Decisions")
    _write_rows(decisions, _research_decision_rows())

    # Every canonical reporting table is rolled into the one cumulative workbook. The per-run
    # Parquet files remain internal reproducibility evidence; users only need to upload this
    # workbook plus MASTER_JSON_PATH for review.
    aggregate_specs = (
        ("Summary History", "summary"),
        ("Actual Trades History", "actual_trades"),
        ("Shadow Trades History", "shadow_trades"),
        ("Request Outcomes", "request_outcomes"),
        ("Equity Curve History", "equity_curve"),
        ("Daily History", "daily_performance"),
        ("Actual Strategy History", "actual_strategy_breakdown"),
        ("Shadow Strategy History", "shadow_strategy_breakdown"),
        ("Symbol History", "symbol_breakdown"),
        ("Actual Cost Breakdown", "actual_cost_breakdown"),
        ("Shadow Cost Breakdown", "shadow_cost_breakdown"),
        ("Actual Exit History", "actual_exit_reason_breakdown"),
        ("Shadow Exit History", "shadow_exit_reason_breakdown"),
        ("Cumulative PnL", "cumulative_pnl"),
        ("Monthly History", "monthly_performance"),
        ("Side History", "side_performance"),
        ("Time Of Day History", "time_of_day_performance"),
        ("Holding Time", "holding_time_distribution"),
        ("Rolling Metrics", "rolling_trade_metrics"),
        ("Diagnostics History", "trade_diagnostics"),
        ("Cost Impact", "cost_impact"),
        ("Outcome Funnel", "outcome_funnel"),
        ("Actual Shadow Compare", "actual_shadow_comparison"),
        ("Provenance History", "provenance"),
    )
    for sheet_name, table_name in aggregate_specs:
        rows: list[dict[str, object]] = []
        for attempt, (report, manifest, _) in enumerate(artifacts, start=1):
            rows.extend(
                _table_rows(
                    report,
                    table_name,
                    attempt=attempt,
                    phase=str(manifest.get("phase")),
                    window_id=report.provenance.window_id,
                )
            )
        _write_rows(workbook.create_sheet(sheet_name), rows)

    latest_report, latest_manifest, _ = artifacts[-1]
    p = latest_report.performance
    latest_values = [
        ("Latest attempt", len(artifacts)),
        ("Phase", latest_manifest.get("phase")),
        ("Window", latest_report.provenance.window_id or "development"),
        ("Strategy version", latest_manifest.get("strategy_version")),
        ("Actual trades", p.actual_trade_count),
        ("Ending capital", float(p.ending_capital)),
        ("Net profit", float(p.net_profit)),
        ("CAGR", float(p.cagr) if p.cagr is not None else None),
        ("Win rate", float(p.win_rate) if p.win_rate is not None else None),
        ("Net profit factor", profit_factor_text(latest_report)),
        (
            "Average net return/trade",
            float(p.average_net_return_per_trade)
            if p.average_net_return_per_trade is not None
            else None,
        ),
        ("Maximum realized DD", float(p.maximum_realized_drawdown_pct)),
        ("Trades/day", float(p.actual_trades_per_day)),
        ("Total costs", float(p.total_costs)),
        (
            "Hard quantitative gates",
            "PASS"
            if latest_report.acceptance.hard_quantitative_targets_pass
            else "FAIL",
        ),
    ]

    dashboard["A1"] = "Strategy 1 — Cumulative Research Dashboard"
    dashboard["A1"].font = Font(bold=True, size=16)
    dashboard["A3"] = "Latest run"
    dashboard["A3"].font = Font(bold=True)
    for row_index, (label, value) in enumerate(latest_values, start=4):
        dashboard.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        dashboard.cell(row=row_index, column=2, value=_excel_value(value))

    gate_start = 4
    dashboard["D3"] = "Latest gate status"
    dashboard["D3"].font = Font(bold=True)
    latest_gate = _gate_history_rows(artifacts)[-1]
    gate_keys = (
        "causality_pass",
        "full_universe_policy",
        "references_excluded",
        "cagr_gt_20pct",
        "win_rate_gt_50pct",
        "net_pf_gt_2",
        "avg_net_return_gt_0_5pct",
        "portfolio_frequency_ge_2_per_day",
        "hard_quantitative_targets_pass",
    )
    for offset, key in enumerate(gate_keys):
        row = gate_start + offset
        dashboard.cell(row=row, column=4, value=key).font = Font(bold=True)
        dashboard.cell(
            row=row,
            column=5,
            value="PASS" if bool(latest_gate.get(key)) else "FAIL",
        )

    # Conditional formatting makes gate failures immediately visible without changing source truth.
    dashboard.conditional_formatting.add(
        f"E{gate_start}:E{gate_start + len(gate_keys) - 1}",
        CellIsRule(operator="equal", formula=['"FAIL"'], stopIfTrue=False),
    )

    _add_dashboard_chart(
        dashboard,
        history,
        title="Ending Capital by Attempt",
        column_header="ending_capital",
        anchor="A22",
    )
    _add_dashboard_chart(
        dashboard,
        history,
        title="Net P&L by Attempt",
        column_header="net_profit",
        anchor="N22",
        chart_type="bar",
    )
    _add_dashboard_chart(
        dashboard,
        history,
        title="Win Rate by Attempt",
        column_header="win_rate",
        anchor="A38",
    )
    _add_dashboard_chart(
        dashboard,
        history,
        title="Profit Factor by Attempt",
        column_header="profit_factor",
        anchor="N38",
    )
    _add_dashboard_chart(
        dashboard,
        history,
        title="Average Net Return by Attempt",
        column_header="average_net_return",
        anchor="A54",
    )
    _add_dashboard_chart(
        dashboard,
        history,
        title="Max Drawdown by Attempt",
        column_header="max_drawdown_pct",
        anchor="N54",
    )
    _add_dashboard_chart(
        dashboard,
        history,
        title="Trade Frequency by Attempt",
        column_header="trades_per_day",
        anchor="A70",
    )
    _add_dashboard_chart(
        dashboard,
        history,
        title="Total Costs by Attempt",
        column_header="total_costs",
        anchor="N70",
        chart_type="bar",
    )

    dashboard.column_dimensions["A"].width = 32
    dashboard.column_dimensions["B"].width = 24
    dashboard.column_dimensions["D"].width = 38
    dashboard.column_dimensions["E"].width = 14
    dashboard.freeze_panes = "A3"

    MASTER_EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    temp_path = MASTER_EXCEL_PATH.with_suffix(".tmp.xlsx")
    workbook.save(temp_path)
    os.replace(temp_path, MASTER_EXCEL_PATH)


def _save_master_plot(
    path: Path,
    title: str,
    x: list[int],
    y: list[float | None],
    ylabel: str,
) -> None:
    fig, axis = plt.subplots()
    valid = [(a, b) for a, b in zip(x, y, strict=True) if b is not None]
    if valid:
        axis.plot([a for a, _ in valid], [b for _, b in valid], marker="o")
        axis.set_xlabel("Research attempt")
        axis.set_ylabel(ylabel)
    else:
        axis.text(0.5, 0.5, "No finite observations", ha="center", va="center")
        axis.set_axis_off()
    axis.set_title(title)
    fig.tight_layout()
    fig.savefig(path, dpi=160)
    plt.close(fig)


def rebuild_master_visuals() -> None:
    artifacts = _run_artifacts()
    if not artifacts:
        return
    MASTER_VISUAL_DIR.mkdir(parents=True, exist_ok=True)
    rows = _run_history_rows(artifacts)
    attempts = [int(row["attempt"]) for row in rows]

    series = (
        ("research_ending_capital.png", "Ending Capital by Attempt", "ending_capital"),
        ("research_net_profit.png", "Net P&L by Attempt", "net_profit"),
        ("research_win_rate.png", "Win Rate by Attempt", "win_rate"),
        ("research_profit_factor.png", "Net Profit Factor by Attempt", "profit_factor"),
        (
            "research_average_net_return.png",
            "Average Net Return per Trade by Attempt",
            "average_net_return",
        ),
        (
            "research_max_drawdown.png",
            "Maximum Realized Drawdown by Attempt",
            "max_drawdown_pct",
        ),
        (
            "research_trade_frequency.png",
            "Actual Trades per Day by Attempt",
            "trades_per_day",
        ),
        ("research_costs.png", "Total Actual Costs by Attempt", "total_costs"),
    )
    for filename, title, field in series:
        values = [
            None if row[field] is None else float(row[field])  # type: ignore[arg-type]
            for row in rows
        ]
        _save_master_plot(
            MASTER_VISUAL_DIR / filename,
            title,
            attempts,
            values,
            field,
        )


def plan_from_registry() -> OOSPlan | None:
    if not OOS_REGISTRY_PATH.exists():
        return None
    with OOSRegistry(OOS_REGISTRY_PATH) as registry:
        try:
            return registry.get_plan(RESEARCH_SCOPE_ID, PLAN_ID)
        except LookupError:
            return None


def print_gate(name: str, passed: bool, detail: str = "") -> None:
    status = "PASS" if passed else "FAIL"
    suffix = f" | {detail}" if detail else ""
    print(f"{status:4}  {name}{suffix}")


def print_research_gates(
    *,
    report: ReportBundle,
    causality: StrategyCausalityReport,
    scanned_symbols: tuple[str, ...],
) -> None:
    print("\n=== RESEARCH GATE SCORECARD ===")
    print_gate(
        "Strategy causality / prefix invariance",
        True,
        f"gate={causality.gate_version}, prefixes={causality.tested_prefix_count}",
    )
    print_gate(
        "Full dynamic historical equity-universe policy",
        bool(scanned_symbols),
        f"scanned={len(scanned_symbols)}",
    )
    print_gate(
        "Reference series excluded from trading universe",
        all(
            symbol not in scanned_symbols
            for symbol in STANDARD_OOS_PARTITION_POLICY.excluded_reference_symbols
        ),
        ",".join(STANDARD_OOS_PARTITION_POLICY.excluded_reference_symbols),
    )
    print_gate("CAGR > 20%", report.acceptance.cagr_pass)
    print_gate("Win rate > 50%", report.acceptance.win_rate_pass)
    print_gate("Net profit factor > 2", report.acceptance.profit_factor_pass)
    print_gate(
        "Average net return/trade > 0.5%",
        report.acceptance.average_net_return_pass,
    )
    print_gate(
        "Portfolio frequency target >= 2 trades/day",
        report.acceptance.frequency_target_met,
        "ideal portfolio target; not part of hard quantitative pass",
    )
    print_gate(
        "Hard quantitative strategy gates",
        report.acceptance.hard_quantitative_targets_pass,
    )


def print_report_summary(report: ReportBundle) -> None:
    p = report.performance
    print("\n=== BACKTEST RESULT ===")
    print(f"Run ID: {report.provenance.run_id}")
    print(f"Actual trades: {p.actual_trade_count}")
    print(f"Shadow trades: {report.shadow_metrics.shadow_trade_count}")
    print(f"Ending capital: {p.ending_capital}")
    print(f"Net profit: {p.net_profit}")
    print(f"Total return: {p.total_return}")
    print(f"CAGR: {p.cagr}")
    print(f"Win rate: {p.win_rate}")
    print(f"Net PF: {profit_factor_text(report)}")
    print(f"Average net return/trade: {p.average_net_return_per_trade}")
    print(f"Maximum realized drawdown: {p.maximum_realized_drawdown_pct}")
    print(f"Actual trades/day: {p.actual_trades_per_day}")
    print(f"Total actual costs: {p.total_costs}")


def run_window(
    *,
    phase: str,
    window_start: date,
    window_end: date,
    strategy: LiquidityShockReclaimStrategy,
    scorer: TradeScorer,
    scorer_provenance: dict[str, object],
    store: ParquetMarketDataStore,
    coverages: tuple[SymbolCoverage, ...],
    data_start: date,
    data_end: date,
    ready_symbol: str,
    ready_at: datetime,
    causality: StrategyCausalityReport,
    git_commit: str,
    runner_sha256: str,
    runner_git_state: str,
    eol_only_paths: tuple[str, ...],
    brokerage_plan: BrokeragePlan,
    slippage_bps: Decimal,
    oos_registry: OOSRegistry | None = None,
    oos_plan: OOSPlan | None = None,
) -> ReportBundle:
    (
        requests,
        scanned_symbols,
        trading_dates,
        skipped_expensive,
        raw_signal_count,
    ) = generate_requests(
        store=store,
        coverages=coverages,
        strategy=strategy,
        scorer=scorer,
        window_start=window_start,
        window_end=window_end,
    )

    if not trading_dates:
        raise RuntimeError("No market trading dates were found in the scored window.")

    print("\n=== SCAN SUMMARY ===")
    print(f"Historical equities scanned: {len(scanned_symbols)}")
    print(f"Raw strategy signals: {raw_signal_count}")
    print(f"Backtest requests: {len(requests)}")
    print(f"Trading dates: {len(trading_dates)}")
    print(f"Skipped because one share exceeded ₹100k notional ceiling: {skipped_expensive}")

    captured_at = datetime.now(IST).replace(microsecond=0)
    margin_snapshot = capture_or_expand_margin_snapshot(
        requests=requests,
        captured_at=captured_at,
    )

    data_manifest, data_subset_fingerprint = build_market_data_manifest(
        tuple(DATA_PATH / f"{symbol}.parquet" for symbol in scanned_symbols),
        dataset_root=DATA_PATH,
        cache_path=MARKET_DATA_FINGERPRINT_CACHE,
    )
    strategy_config_fingerprint = canonical_fingerprint(dict(strategy.parameters))
    margin_snapshot_fingerprint = (
        canonical_fingerprint(
            margin_snapshot.model_dump(mode="json", exclude_computed_fields=True)
        )
        if margin_snapshot is not None
        else None
    )
    scorer_kind = str(scorer_provenance["scorer_kind"])
    reproducibility: dict[str, object] = {
        "data_subset_fingerprint": data_subset_fingerprint,
        "data_manifest": [asdict(record) for record in data_manifest],
        "margin_snapshot_fingerprint": margin_snapshot_fingerprint,
        "margin_snapshot_id": (
            margin_snapshot.snapshot_id if margin_snapshot is not None else None
        ),
        "ml_scorer_kind": scorer_kind,
        "ml_model_version": (
            BOOTSTRAP_MODEL_VERSION
            if scorer_kind == "BOOTSTRAP"
            else getattr(getattr(scorer, "metadata", None), "model_version", "UNKNOWN")
        ),
        "ml_artifact_fingerprint": scorer_provenance.get("artifact_fingerprint"),
        "ml_feature_schema": STRATEGY1_META_FEATURE_SCHEMA.model_dump(mode="json"),
        "ml_training_data_lineage_fingerprint": scorer_provenance.get(
            "training_data_lineage_fingerprint"
        ),
        "ml_evaluation_identity": scorer_provenance.get("evaluation_identity"),
        "ml_selection_reason": scorer_provenance["reason"],
        "strategy_config_fingerprint": strategy_config_fingerprint,
        "cost_policy": f"FIXED_CURRENT:{brokerage_plan.value}",
        "slippage_policy": "ADVERSE_FIXED_BPS_PER_FILL",
        "slippage_bps_per_fill": str(slippage_bps),
        "backtester_version": BACKTESTER_VERSION,
        "reporting_version": REPORTING_VERSION,
        "environment": environment_snapshot(),
        "runner_sha256": runner_sha256,
        "git_commit": git_commit,
        "git_state": runner_git_state,
    }
    reproducibility["run_input_fingerprint"] = canonical_fingerprint(reproducibility)

    run_id = stable_run_id(
        phase=phase,
        strategy_version=strategy.strategy_version,
        window_start=window_start,
        window_end=window_end,
        git_commit=git_commit,
        runner_sha256=runner_sha256,
        brokerage_plan=brokerage_plan,
        slippage_bps=slippage_bps,
        run_input_fingerprint=str(reproducibility["run_input_fingerprint"]),
    )
    current_oos_window = None
    if phase == "oos":
        if oos_registry is None or oos_plan is None:
            raise RuntimeError("OOS run requires the persistent OOS registry and plan.")
        current_oos_window = oos_registry.next_testable_window(RESEARCH_SCOPE_ID, PLAN_ID)
        if current_oos_window is None:
            raise RuntimeError(
                "No OOS window is currently testable. Review/authorize the prior TESTED window "
                "before attempting another OOS run."
            )
        if (
            current_oos_window.start_date != window_start
            or current_oos_window.end_date != window_end
        ):
            raise RuntimeError("Requested OOS window is not the registry's next testable window.")
    run_dir, final_run_dir = prepare_staging_run_directory(
        RESULTS_ROOT / "runs", run_id
    )

    backtester = create_backtester(
        store=store,
        margin_snapshot=margin_snapshot,
        slippage_bps=slippage_bps,
    )
    result = backtester.run(
        BacktestConfig(
            run_id=run_id,
            git_commit=git_commit,
            window_start=as_ist_midnight(window_start),
            window_end=as_ist_midnight(window_end),
            brokerage_plan=brokerage_plan,
            initial_capital=INITIAL_CAPITAL,
            forced_exit_time=FORCED_BACKTEST_EXIT_TIME,
        ),
        requests,
    )

    # Persist the exact immutable backtest result before any OOS state mutation.
    write_json_model(run_dir / "backtest_result.json", result)

    result_fingerprint = fingerprint_backtest_result(result)
    report = build_report(
        result,
        ReportContext(
            report_id=f"{run_id}-report",
            generated_at=captured_at,
            trading_dates=trading_dates,
            research_scope_id=(RESEARCH_SCOPE_ID if current_oos_window else None),
            plan_id=(PLAN_ID if current_oos_window else None),
            window_id=(current_oos_window.window_id if current_oos_window else None),
            oos_result_fingerprint=(result_fingerprint if current_oos_window else None),
        ),
    )

    # Canonical per-run machine-readable derivatives.
    write_json_model(run_dir / "report_bundle.json", report)
    dataset_paths = write_report_dataset(report, run_dir / "parquet_report")

    # Comprehensive Matplotlib diagnostics from the existing reporting layer.
    visual_paths = write_visual_report(report, run_dir / "visuals")

    write_run_manifest(
        path=run_dir / "run_manifest.json",
        phase=phase,
        report=report,
        git_commit=git_commit,
        runner_sha256=runner_sha256,
        runner_git_state=runner_git_state,
        eol_only_paths=eol_only_paths,
        data_start=data_start,
        data_end=data_end,
        ready_symbol=ready_symbol,
        ready_at=ready_at,
        scanned_symbols=scanned_symbols,
        raw_signal_count=raw_signal_count,
        request_count=len(requests),
        trading_day_count=len(trading_dates),
        skipped_expensive=skipped_expensive,
        slippage_bps=slippage_bps,
        margin_snapshot_id=(
            margin_snapshot.snapshot_id if margin_snapshot is not None else None
        ),
        causality=causality,
        reproducibility=reproducibility,
    )

    finalized_run_dir = finalize_staging_run_directory(run_dir, final_run_dir)

    if current_oos_window is not None:
        assert oos_registry is not None
        oos_registry.register_test_result(
            RESEARCH_SCOPE_ID,
            PLAN_ID,
            current_oos_window.window_id,
            result,
            OOSAuditContext(
                event_id=f"strategy1-oos-test-{uuid4().hex}",
                occurred_at=captured_at,
                git_commit=git_commit,
            ),
            tested_strategy_versions=((strategy.strategy_id, strategy.strategy_version),),
            scanned_symbols=scanned_symbols,
        )

    print_report_summary(report)
    print_research_gates(
        report=report,
        causality=causality,
        scanned_symbols=scanned_symbols,
    )
    print(
        f"\nPer-run Parquet tables: {len(dataset_paths)} -> "
        f"{finalized_run_dir / 'parquet_report'}"
    )
    print(
        f"Per-run Matplotlib figures: {len(visual_paths)} -> "
        f"{finalized_run_dir / 'visuals'}"
    )
    return report


def authorize_oos(
    *,
    registry: OOSRegistry,
    git_commit: str,
    requested_window_id: str | None,
) -> OOSPlan:
    plan = registry.get_plan(RESEARCH_SCOPE_ID, PLAN_ID)
    tested = [
        window
        for window in plan.oos_windows
        if window.state is OOSWindowState.TESTED
    ]
    if requested_window_id is not None:
        tested = [window for window in tested if window.window_id == requested_window_id]
    if len(tested) != 1:
        raise RuntimeError(
            "authorize-oos requires exactly one TESTED window. "
            f"Matching TESTED windows: {[window.window_id for window in tested]}"
        )
    window = tested[0]
    now = datetime.now(IST).replace(microsecond=0)
    registry.mark_consumed(
        RESEARCH_SCOPE_ID,
        PLAN_ID,
        window.window_id,
        OOSAuditContext(
            event_id=f"strategy1-oos-consume-{uuid4().hex}",
            occurred_at=now,
            git_commit=git_commit,
        ),
    )
    registry.authorize_training(
        RESEARCH_SCOPE_ID,
        PLAN_ID,
        window.window_id,
        OOSAuditContext(
            event_id=f"strategy1-oos-train-{uuid4().hex}",
            occurred_at=datetime.now(IST).replace(microsecond=0),
            git_commit=git_commit,
        ),
    )
    updated = registry.get_plan(RESEARCH_SCOPE_ID, PLAN_ID)
    print(
        f"OOS {window.window_id}: TESTED -> CONSUMED -> TRAINING_ALLOWED. "
        "The next chronological OOS window may now be tested."
    )
    return updated


def main() -> None:
    args = parse_args()
    try:
        slippage_bps = Decimal(str(args.slippage_bps))
    except Exception as error:
        raise ValueError("--slippage-bps must be numeric") from error
    if not slippage_bps.is_finite() or slippage_bps < 0:
        raise ValueError("--slippage-bps must be finite and non-negative")
    brokerage_plan = BrokeragePlan(args.brokerage_plan)

    (
        git_commit,
        runner_sha256,
        eol_only_paths,
        runner_git_state,
    ) = repository_reproducibility_state()

    print("=== STRATEGY 1 RESEARCH PREFLIGHT ===")
    print(f"Mode: {args.mode}")
    print(f"Git HEAD: {git_commit}")
    print(f"Runner SHA-256: {runner_sha256}")
    print(f"Runner Git state: {runner_git_state}")
    if eol_only_paths:
        print(f"Allowed line-ending-only tracked changes: {', '.join(eol_only_paths)}")
    print(f"Brokerage plan: {brokerage_plan.value}")
    print(f"Slippage: {slippage_bps} bps adverse per fill")

    RESULTS_ROOT.mkdir(parents=True, exist_ok=True)

    if args.mode == "authorize-oos":
        if not OOS_REGISTRY_PATH.exists():
            raise RuntimeError("No frozen OOS registry exists yet.")
        with OOSRegistry(OOS_REGISTRY_PATH) as registry:
            updated_plan = authorize_oos(
                registry=registry,
                git_commit=git_commit,
                requested_window_id=args.window_id,
            )
        json_path, workbook_path = build_review_artifacts(
            results_root=RESULTS_ROOT,
            research_scope_id=RESEARCH_SCOPE_ID,
            decisions=RESEARCH_DECISION_HISTORY,
            oos_plan=updated_plan,
            generated_at=datetime.now(IST).replace(microsecond=0),
        )
        print(f"Cumulative research JSON updated: {json_path}")
        print(f"Master Excel updated: {workbook_path}")
        return

    strategy = LiquidityShockReclaimStrategy()
    strategy_config_fingerprint = canonical_fingerprint(dict(strategy.parameters))
    completed_discovery = discover_research_runs(RESULTS_ROOT)
    model_selection = select_scope_trade_scorer(
        model_root=MODEL_ROOT,
        bootstrap_model_version=BOOTSTRAP_MODEL_VERSION,
        research_scope_id=RESEARCH_SCOPE_ID,
        plan_id=PLAN_ID,
        strategy_id=strategy.strategy_id,
        feature_schema=STRATEGY1_META_FEATURE_SCHEMA,
        strategy_config_fingerprint=strategy_config_fingerprint,
        completed_results=tuple(run.result for run in completed_discovery.completed),
    )
    scorer = model_selection.scorer
    scorer_provenance: dict[str, object] = {
        "scorer_kind": model_selection.scorer_kind,
        "reason": model_selection.reason,
        "artifact_fingerprint": (
            model_selection.artifact_identity.artifact_fingerprint
            if model_selection.artifact_identity is not None
            else None
        ),
        "training_data_lineage_fingerprint": (
            canonical_fingerprint(
                [
                    source.model_dump(mode="json")
                    for source in getattr(
                        getattr(scorer, "metadata", None), "training_sources", ()
                    )
                ]
            )
            if model_selection.scorer_kind == "TRAINED"
            else None
        ),
        "evaluation_identity": (
            model_selection.evaluation.model_dump(mode="json")
            if model_selection.evaluation is not None
            else None
        ),
    }
    print(
        f"ML scorer: {model_selection.scorer_kind} | {model_selection.reason}"
    )
    store = ParquetMarketDataStore(MarketDataConfig(dataset_path=DATA_PATH))
    coverages = tuple(store.get_symbols_coverage())
    data_start, data_end = derive_equity_data_horizon(coverages)
    ready_symbol, ready_at = derive_earliest_strategy_ready(
        coverages=coverages,
        strategy=strategy,
    )
    with OOSRegistry(OOS_REGISTRY_PATH) as preflight_registry:
        _, development_end, _, _ = development_dates(
            data_start=data_start,
            data_end=data_end,
            ready_at=ready_at,
            registry=preflight_registry,
        )
    causality = run_real_causality_gate(
        store=store,
        coverages=coverages,
        strategy=strategy,
        allowed_end_exclusive=development_end,
    )

    print(f"Equity data horizon: {data_start} -> {data_end}")
    print(
        "Earliest causal Strategy-1 readiness: "
        f"{ready_symbol} @ {ready_at.isoformat()}"
    )
    print(
        "CAUSALITY GATE PASS: "
        f"version={causality.gate_version} "
        f"prefixes={causality.tested_prefix_count} "
        f"signals={causality.full_signal_count}"
    )

    with OOSRegistry(OOS_REGISTRY_PATH) as registry:
        (
            development_start,
            development_end,
            proposed_oos_start,
            existing_plan,
        ) = development_dates(
            data_start=data_start,
            data_end=data_end,
            ready_at=ready_at,
            registry=registry,
        )

        if args.mode == "development":
            print(
                "Development scored interval: "
                f"{development_start} -> {development_end}"
            )
            print(
                "Warm-up/training history begins at the earliest equity date: "
                f"{data_start}"
            )
            report = run_window(
                phase="development",
                window_start=development_start,
                window_end=development_end,
                strategy=strategy,
                scorer=scorer,
                scorer_provenance=scorer_provenance,
                store=store,
                coverages=coverages,
                data_start=data_start,
                data_end=data_end,
                ready_symbol=ready_symbol,
                ready_at=ready_at,
                causality=causality,
                git_commit=git_commit,
                runner_sha256=runner_sha256,
                runner_git_state=runner_git_state,
                eol_only_paths=eol_only_paths,
                brokerage_plan=brokerage_plan,
                slippage_bps=slippage_bps,
            )

            # Freeze the standard OOS plan only after a successful development result exists.
            plan = get_or_create_standard_plan(
                registry=registry,
                data_start=data_start,
                data_end=data_end,
                earliest_oos_start=development_end,
                git_commit=git_commit,
                occurred_at=report.provenance.generated_at,
            )

        else:
            plan = existing_plan
            if plan is None:
                if not discover_research_runs(RESULTS_ROOT).completed:
                    raise RuntimeError(
                        "Run the development backtest successfully before creating/testing OOS."
                    )
                plan = get_or_create_standard_plan(
                    registry=registry,
                    data_start=data_start,
                    data_end=data_end,
                    earliest_oos_start=proposed_oos_start,
                    git_commit=git_commit,
                    occurred_at=datetime.now(IST).replace(microsecond=0),
                )

            current = registry.next_testable_window(RESEARCH_SCOPE_ID, PLAN_ID)
            if current is None:
                states = {
                    window.window_id: window.state.value for window in plan.oos_windows
                }
                raise RuntimeError(
                    "No OOS window is currently testable. "
                    "If the last OOS is TESTED, review it first and then run "
                    "--mode authorize-oos. States: "
                    + json.dumps(states, sort_keys=True)
                )
            print(
                f"Next governed OOS window: {current.window_id} "
                f"{current.start_date} -> {current.end_date}"
            )
            report = run_window(
                phase="oos",
                window_start=current.start_date,
                window_end=current.end_date,
                strategy=strategy,
                scorer=scorer,
                scorer_provenance=scorer_provenance,
                store=store,
                coverages=coverages,
                data_start=data_start,
                data_end=data_end,
                ready_symbol=ready_symbol,
                ready_at=ready_at,
                causality=causality,
                git_commit=git_commit,
                runner_sha256=runner_sha256,
                runner_git_state=runner_git_state,
                eol_only_paths=eol_only_paths,
                brokerage_plan=brokerage_plan,
                slippage_bps=slippage_bps,
                oos_registry=registry,
                oos_plan=plan,
            )
            plan = registry.get_plan(RESEARCH_SCOPE_ID, PLAN_ID)

    master_json_path, master_workbook_path = build_review_artifacts(
        results_root=RESULTS_ROOT,
        research_scope_id=RESEARCH_SCOPE_ID,
        decisions=RESEARCH_DECISION_HISTORY,
        oos_plan=plan,
        generated_at=datetime.now(IST).replace(microsecond=0),
    )
    print("\n=== ARTIFACT SUMMARY ===")
    print("UPLOAD ONLY THESE TWO FILES FOR REVIEW:")
    print(f"  JSON:  {master_json_path}")
    print(f"  Excel: {master_workbook_path}")
    print(
        "Per-run Parquet/JSON artifacts and Matplotlib diagnostics remain internal reproducibility "
        "evidence. They do not need to be uploaded for routine research review."
    )

    if args.mode == "development":
        print(
            "\nNEXT: review the development result and master workbook. "
            "When approved, run:\n"
            "  python scripts\\run_strategy1_development_backtest.py --mode next-oos"
        )
    else:
        print(
            "\nSTOP: this OOS window is now TESTED, not yet training-authorized. "
            "Review its result before running:\n"
            f"  python scripts\\run_strategy1_development_backtest.py "
            f"--mode authorize-oos --window-id {report.provenance.window_id}"
        )


if __name__ == "__main__":
    main()
