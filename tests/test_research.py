from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from zoneinfo import ZoneInfo

import polars as pl
import pytest
from openpyxl import load_workbook

from algo_trader.backtest import BacktestRunResult
from algo_trader.costs import BrokeragePlan
from algo_trader.domain import MLScore, Side, Signal
from algo_trader.reporting import ReportContext, build_report
from algo_trader.research import (
    ResearchArtifactIntegrityError,
    ResearchDecisionRecord,
    ResearchStrategySpec,
    build_market_data_manifest,
    build_review_artifacts,
    discover_research_runs,
    finalize_staging_run_directory,
    prepare_staging_run_directory,
    score_and_build_requests,
)

IST = ZoneInfo("Asia/Kolkata")
NOW = datetime(2026, 8, 15, 12, tzinfo=IST)


def result(run_id: str = "run-1") -> BacktestRunResult:
    return BacktestRunResult(
        run_id=run_id,
        git_commit="commit-1",
        backtester_version="1",
        window_start=datetime(2025, 1, 1, tzinfo=IST),
        window_end=datetime(2025, 2, 1, tzinfo=IST),
        cost_policy_id="cost-policy",
        cost_policy_source_as_of_date=date(2026, 8, 15),
        brokerage_plan=BrokeragePlan.PLUS,
        starting_capital=Decimal("100000"),
        ending_capital=Decimal("100000"),
        capital_exhausted=False,
        actual_trade_records=(),
        shadow_trade_records=(),
        request_results=(),
        ending_portfolio_state=None,
        symbols=(),
        strategy_versions=(),
        ml_model_versions=(),
    )


def completed_run(results_root: Path, run_id: str = "run-1") -> Path:
    directory = results_root / "runs" / run_id
    directory.mkdir(parents=True)
    selected = result(run_id)
    report = build_report(
        selected,
        ReportContext(
            report_id=f"{run_id}-report",
            generated_at=NOW,
            trading_dates=(date(2025, 1, 2),),
        ),
    )
    (directory / "backtest_result.json").write_text(
        selected.model_dump_json(), encoding="utf-8"
    )
    (directory / "report_bundle.json").write_text(
        report.model_dump_json(), encoding="utf-8"
    )
    (directory / "run_manifest.json").write_text(
        json.dumps(
            {
                "phase": "development",
                "completed_at": NOW.isoformat(),
                "run_input_fingerprint": "input-sha",
                "causality_gate": {"pass": True},
            }
        ),
        encoding="utf-8",
    )
    return directory


def decision(run_id: str = "run-1") -> ResearchDecisionRecord:
    return ResearchDecisionRecord(
        decision_id="decision-1",
        source_run_ids=(run_id,),
        strategy_id="strategy",
        strategy_version="1",
        research_scope_id="scope",
        decision="REVIEWED",
        diagnosis="Synthetic fixture decision.",
        changes_authorized=(),
        changes_rejected=("parameter fishing",),
        next_action="Retain the governed workflow.",
        recorded_at=NOW,
        git_commit="commit-1",
    )


def test_completed_run_builds_exact_valid_two_file_review_bundle(tmp_path: Path) -> None:
    completed_run(tmp_path)

    json_path, workbook_path = build_review_artifacts(
        results_root=tmp_path,
        research_scope_id="scope",
        decisions=(decision(),),
        oos_plan={"window": "development"},
        generated_at=NOW,
    )

    review = tmp_path / "UPLOAD_FOR_REVIEW"
    assert tuple(sorted(path.name for path in review.iterdir())) == (
        "strategy1_research_history.json",
        "strategy1_research_master.xlsx",
    )
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["runs"][0]["backtest_result"]["run_id"] == "run-1"
    assert payload["runs"][0]["report_bundle"]["provenance"]["run_id"] == "run-1"
    assert payload["runs"][0]["run_manifest"]["phase"] == "development"
    assert set(payload["runs"][0]["source_sha256"]) == {
        "backtest_result.json",
        "report_bundle.json",
        "run_manifest.json",
    }
    workbook = load_workbook(workbook_path, read_only=True)
    assert "Signal Features" in workbook.sheetnames
    workbook.close()


def test_incomplete_run_is_skipped_but_corrupt_completed_run_fails(tmp_path: Path) -> None:
    incomplete = tmp_path / "runs" / "partial"
    incomplete.mkdir(parents=True)
    (incomplete / "backtest_result.json").write_text("{}", encoding="utf-8")
    discovery = discover_research_runs(tmp_path)
    assert discovery.completed == ()
    assert discovery.skipped_incomplete[0][0] == incomplete

    corrupt = completed_run(tmp_path, "corrupt")
    (corrupt / "report_bundle.json").write_text("not-json", encoding="utf-8")
    with pytest.raises(ResearchArtifactIntegrityError, match="corrupt"):
        discover_research_runs(tmp_path)


def test_atomic_run_finalization_requires_all_canonical_artifacts(tmp_path: Path) -> None:
    staging, final = prepare_staging_run_directory(tmp_path / "runs", "run-1")
    with pytest.raises(ResearchArtifactIntegrityError, match="missing"):
        finalize_staging_run_directory(staging, final)
    assert staging.is_dir()
    assert not final.exists()

    source = completed_run(tmp_path / "source")
    for name in ("backtest_result.json", "report_bundle.json", "run_manifest.json"):
        (staging / name).write_bytes((source / name).read_bytes())
    assert finalize_staging_run_directory(staging, final) == final
    assert final.is_dir()
    assert not staging.exists()


def test_review_generation_failure_never_exposes_success_directory(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from algo_trader.research import artifacts

    completed_run(tmp_path)
    monkeypatch.setattr(
        artifacts,
        "_write_review_workbook",
        lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("write failed")),
    )

    with pytest.raises(RuntimeError, match="write failed"):
        build_review_artifacts(
            results_root=tmp_path,
            research_scope_id="scope",
            decisions=(decision(),),
            oos_plan=None,
            generated_at=NOW,
        )
    assert not (tmp_path / "UPLOAD_FOR_REVIEW").exists()
    assert not (tmp_path / "UPLOAD_FOR_REVIEW.tmp").exists()


def test_legacy_computed_cost_totals_are_ignored_only_in_memory() -> None:
    from algo_trader.research.artifacts import _remove_legacy_computed_cost_totals

    payload = {
        "trade": {
            "round_trip_cost_breakdown": {
                "total": "10",
                "entry": {"total": "4", "brokerage": "1"},
                "exit": {"total": "6", "brokerage": "2"},
            }
        },
        "unrelated_total": "preserved",
    }
    _remove_legacy_computed_cost_totals(payload)
    costs = payload["trade"]["round_trip_cost_breakdown"]
    assert costs == {
        "entry": {"brokerage": "1"},
        "exit": {"brokerage": "2"},
    }
    assert payload["unrelated_total"] == "preserved"


def test_generic_research_orchestration_scores_before_request_sizing() -> None:
    events: list[str] = []
    signal = Signal(
        strategy_id="fake",
        strategy_version="1",
        symbol="AAA",
        timestamp=NOW,
        side=Side.LONG,
    )

    class Scorer:
        def score(self, supplied: Signal) -> MLScore:
            assert supplied is signal
            events.append("score")
            return MLScore(
                model_version="trained-v1",
                quality_score=0.8,
                calibrated_probability=0.8,
                predicted_net_return=0.005,
                recommended_notional=90_000,
            )

    def builder(supplied: Signal, score: MLScore) -> int:
        assert supplied is signal
        events.append("quantity")
        return score.recommended_notional // 100

    assert score_and_build_requests((signal,), Scorer(), builder) == (900,)
    assert events == ["score", "quantity"]


def test_generic_research_spec_accepts_structural_strategy_without_inheritance() -> None:
    class FakeStrategy:
        strategy_id = "fake"
        strategy_version = "1"
        parameters = {"threshold": 1}
        warmup_bars = 2

        def generate_signals(self, candles: pl.DataFrame) -> list[Signal]:
            del candles
            return []

    spec = ResearchStrategySpec(
        research_scope_id="fake-scope",
        plan_id="fake-plan",
        output_slug="fake-output",
        strategy_factory=FakeStrategy,
        request_builder=lambda signal, score: (signal, score),
    )

    strategy = spec.create_strategy()
    assert strategy.strategy_id == "fake"
    assert strategy.generate_signals(pl.DataFrame()) == []


def test_market_data_manifest_cache_preserves_strong_identity(tmp_path: Path) -> None:
    dataset = tmp_path / "data"
    dataset.mkdir()
    path = dataset / "AAA.parquet"
    pl.DataFrame(
        {
            "timestamp": [NOW, NOW.replace(minute=5)],
            "open": [100.0, 101.0],
            "high": [101.0, 102.0],
            "low": [99.0, 100.0],
            "close": [100.5, 101.5],
            "volume": [10.0, 20.0],
            "symbol": ["AAA", "AAA"],
        }
    ).write_parquet(path)
    cache = tmp_path / "cache.json"
    first, first_fingerprint = build_market_data_manifest(
        (path,), dataset_root=dataset, cache_path=cache
    )
    second, second_fingerprint = build_market_data_manifest(
        (path,), dataset_root=dataset, cache_path=cache
    )
    assert first == second
    assert first_fingerprint == second_fingerprint
    assert first[0].row_count == 2
    assert len(first[0].sha256) == 64
