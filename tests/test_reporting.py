from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import polars as pl
import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from algo_trader.backtest import (
    BacktestRequestOutcome,
    BacktestRequestResult,
    BacktestRunResult,
    BacktestTradeRecord,
    BacktestTradeRequest,
)
from algo_trader.costs import BrokeragePlan, LegCostBreakdown, RoundTripCostBreakdown
from algo_trader.domain import (
    ExitReason,
    Fill,
    MLScore,
    OrderIntent,
    Side,
    Signal,
    SignalStatus,
    Trade,
)
from algo_trader.oos import OOSAuditContext, OOSTestRecord, fingerprint_backtest_result
from algo_trader.portfolio import (
    AllocationCandidate,
    AllocationDecision,
    AllocationOutcome,
    CapitalReservation,
    MarginRequirementQuote,
)
from algo_trader.reporting import (
    COMPARISON_VISUAL_FILENAMES,
    REPORT_TABLE_FILENAMES,
    REPORTING_VERSION,
    ReportContext,
    ReportingIntegrityError,
    build_report,
    build_report_comparison,
    report_tables,
    write_comparison_excel_report,
    write_comparison_visual_report,
    write_excel_report,
    write_report_dataset,
    write_visual_report,
)

MARKET_TIMEZONE = ZoneInfo("Asia/Kolkata")
ZERO = Decimal("0")


def at(day: int, hour: int, minute: int = 0) -> datetime:
    return datetime(2025, 1, day, hour, minute, tzinfo=MARKET_TIMEZONE)


def leg(turnover: Decimal, factor: Decimal = Decimal("1")) -> LegCostBreakdown:
    return LegCostBreakdown(
        turnover=turnover,
        brokerage=Decimal("1") * factor,
        exchange_transaction_charge=Decimal("1.5") * factor,
        sebi_turnover_fee=Decimal("2") * factor,
        ipft=Decimal("2.5") * factor,
        stt=Decimal("3") * factor,
        stamp_duty=Decimal("3.5") * factor,
        gst=Decimal("4") * factor,
    )


def make_record(
    name: str,
    net_pnl: Decimal,
    exit_at: datetime,
    *,
    shadow: bool = False,
    side: Side = Side.LONG,
    symbol: str = "AAA",
    quality: float = 0.7,
    exit_reason: ExitReason = ExitReason.TIME_EXIT,
) -> BacktestTradeRecord:
    generated = Signal(
        strategy_id=name,
        strategy_version="1",
        symbol=symbol,
        timestamp=exit_at - timedelta(hours=1),
        side=side,
    )
    order = OrderIntent(
        signal=generated,
        timestamp=exit_at - timedelta(minutes=55),
        quantity=500,
        requested_notional=50_000,
    )
    score = MLScore(
        model_version="model-1",
        quality_score=quality,
        calibrated_probability=0.6,
        predicted_net_return=0.01,
        recommended_notional=50_000,
    )
    candidate = AllocationCandidate(order_intent=order, ml_score=score)
    quote = MarginRequirementQuote(provider_id="margin-v1", required_margin=Decimal("10000"))
    if shadow:
        decision = AllocationDecision(
            candidate=candidate,
            outcome=AllocationOutcome.CAPACITY_REJECTED,
            margin_quote=quote,
            signal=generated.model_copy(update={"status": SignalStatus.CAPACITY_REJECTED}),
        )
        trade_signal = decision.signal
    else:
        reservation = CapitalReservation(candidate=candidate, margin_quote=quote)
        decision = AllocationDecision(
            candidate=candidate,
            outcome=AllocationOutcome.ALLOCATED,
            margin_quote=quote,
            signal=generated,
            reservation=reservation,
        )
        trade_signal = generated.model_copy(update={"status": SignalStatus.EXECUTED})
    breakdown = RoundTripCostBreakdown(
        schedule_id="cost-policy",
        entry=leg(Decimal("50000"), Decimal("0.4")),
        exit=leg(Decimal("50500"), Decimal("0.6")),
    )
    trade = Trade(
        signal=trade_signal,
        ml_score=score,
        target_notional=50_000,
        entry_fill=Fill(
            timestamp=exit_at - timedelta(minutes=50),
            price=Decimal("100"),
            quantity=500,
            is_simulated=True,
        ),
        exit_fill=Fill(
            timestamp=exit_at,
            price=Decimal("101"),
            quantity=500,
            is_simulated=True,
        ),
        gross_pnl=net_pnl + breakdown.total,
        total_costs=breakdown.total,
        net_pnl=net_pnl,
        mfe_return=Decimal("0.02"),
        mae_return=Decimal("-0.01"),
        exit_reason=exit_reason,
        is_shadow=shadow,
    )
    return BacktestTradeRecord(
        trade=trade,
        round_trip_cost_breakdown=breakdown,
        cost_policy_id="cost-policy",
        allocation_identity=candidate.identity,
        allocation_decision=decision,
    )


def make_result(
    *,
    actual: tuple[BacktestTradeRecord, ...] | None = None,
    shadow: tuple[BacktestTradeRecord, ...] | None = None,
    ending_capital: Decimal | None = None,
) -> BacktestRunResult:
    actual_records = actual if actual is not None else (
        make_record("beta", Decimal("500"), at(2, 10), symbol="BBB"),
        make_record(
            "alpha",
            Decimal("-200"),
            at(3, 10),
            side=Side.SHORT,
            exit_reason=ExitReason.STOP_LOSS,
        ),
        make_record("alpha", ZERO, at(3, 10), exit_reason=ExitReason.MANUAL),
    )
    shadow_records = shadow if shadow is not None else (
        make_record("beta", Decimal("1000"), at(4, 10), shadow=True),
    )
    expected_ending = Decimal("100000") + sum(
        (record.trade.net_pnl for record in actual_records), start=ZERO
    )
    return BacktestRunResult(
        run_id="run-1",
        git_commit="deadbeef",
        backtester_version="1",
        window_start=at(1, 9, 15),
        window_end=at(5, 15, 30),
        cost_policy_id="cost-policy",
        cost_policy_source_as_of_date=date(2026, 8, 14),
        brokerage_plan=BrokeragePlan.PLUS,
        starting_capital=Decimal("100000"),
        ending_capital=expected_ending if ending_capital is None else ending_capital,
        capital_exhausted=False,
        actual_trade_records=actual_records,
        shadow_trade_records=shadow_records,
        request_results=(),
        ending_portfolio_state=None,
        symbols=("BBB", "AAA"),
        strategy_versions=(("beta", "1"), ("alpha", "1")),
        ml_model_versions=("model-1",),
    )


def context() -> ReportContext:
    return ReportContext(
        report_id="report-1",
        generated_at=datetime(2026, 8, 14, 12, tzinfo=MARKET_TIMEZONE),
        trading_dates=tuple(date(2025, 1, day) for day in range(1, 6)),
    )


def test_report_context_requires_explicit_aware_time_and_valid_trading_dates() -> None:
    values = context().model_dump()
    for update, message in (
        ({"generated_at": datetime(2026, 8, 14, 12)}, "timezone-aware"),
        ({"trading_dates": ()}, "at least one"),
        ({"trading_dates": (date(2025, 1, 2), date(2025, 1, 1))}, "chronological"),
        ({"trading_dates": (date(2025, 1, 1), date(2025, 1, 1))}, "duplicates"),
        ({"trading_dates": (at(1, 9),)}, "not datetime"),
    ):
        with pytest.raises((ValidationError, TypeError), match=message):
            ReportContext(**(values | update))


def test_trading_dates_must_intersect_half_open_backtest_window() -> None:
    base = make_result(actual=(), shadow=(), ending_capital=Decimal("100000"))
    partial = base.model_copy(update={"window_start": at(2, 12), "window_end": at(3, 0)})
    assert build_report(
        partial,
        context().model_copy(update={"trading_dates": (date(2025, 1, 2),)}),
    )
    with pytest.raises(ReportingIntegrityError, match="intersect"):
        build_report(
            partial,
            context().model_copy(update={"trading_dates": (date(2025, 1, 3),)}),
        )


@pytest.mark.parametrize(
    ("window_start", "window_end", "valid_dates", "invalid_dates"),
    [
        (at(1, 0), at(2, 0), (date(2025, 1, 1),), (date(2025, 1, 2),)),
        (
            at(1, 0),
            at(2, 12),
            (date(2025, 1, 1), date(2025, 1, 2)),
            (date(2025, 1, 3),),
        ),
        (at(1, 12), at(2, 0), (date(2025, 1, 1),), (date(2025, 1, 2),)),
    ],
)
def test_half_open_local_day_intersection_examples(
    window_start: datetime,
    window_end: datetime,
    valid_dates: tuple[date, ...],
    invalid_dates: tuple[date, ...],
) -> None:
    result = make_result(actual=(), shadow=(), ending_capital=Decimal("100000")).model_copy(
        update={"window_start": window_start, "window_end": window_end}
    )
    assert build_report(
        result, context().model_copy(update={"trading_dates": valid_dates})
    )
    for invalid in invalid_dates:
        with pytest.raises(ReportingIntegrityError, match="intersect"):
            build_report(
                result,
                context().model_copy(update={"trading_dates": (invalid,)}),
            )


def test_actual_metrics_reconcile_and_exclude_shadow_economics() -> None:
    report = build_report(make_result(), context())
    metrics = report.performance

    assert metrics.net_profit == Decimal("300")
    assert REPORTING_VERSION == "3"
    assert report.provenance.reporting_version == "3"
    assert metrics.ending_capital == Decimal("100300")
    assert metrics.actual_trade_count == 3
    assert (
        metrics.winning_trade_count,
        metrics.losing_trade_count,
        metrics.breakeven_trade_count,
    ) == (1, 1, 1)
    assert metrics.win_rate == Decimal("1") / Decimal("3")
    assert metrics.average_net_return_per_trade == Decimal("0.002")
    assert metrics.median_net_return_per_trade == ZERO
    assert metrics.best_trade_net_pnl == Decimal("500")
    assert metrics.worst_trade_net_pnl == Decimal("-200")
    assert metrics.total_return == Decimal("0.003")
    assert metrics.average_mfe_return == Decimal("0.02")
    assert metrics.average_mae_return == Decimal("-0.01")
    assert report.shadow_metrics.hypothetical_net_pnl == Decimal("1000")
    assert report.shadow_metrics.economic_status.startswith("HYPOTHETICAL")


def test_inconsistent_capital_and_costs_are_rejected_exactly() -> None:
    with pytest.raises(ReportingIntegrityError, match="ending_capital"):
        build_report(make_result(ending_capital=Decimal("99999")), context())

    valid = make_result()
    bad_trade = valid.actual_trade_records[0].trade.model_copy(update={"total_costs": Decimal("0")})
    bad_record = valid.actual_trade_records[0].model_copy(update={"trade": bad_trade})
    bad_result = valid.model_copy(
        update={"actual_trade_records": (bad_record, *valid.actual_trade_records[1:])}
    )
    with pytest.raises(ReportingIntegrityError, match="cost"):
        build_report(bad_result, context())


@pytest.mark.parametrize(
    ("pnl", "value", "unbounded", "undefined"),
    [
        ((Decimal("500"), Decimal("-200")), Decimal("2.5"), False, False),
        ((Decimal("500"),), None, True, False),
        ((Decimal("-200"),), ZERO, False, False),
        ((ZERO,), None, False, True),
    ],
)
def test_profit_factor_semantics(
    pnl: tuple[Decimal, ...],
    value: Decimal | None,
    unbounded: bool,
    undefined: bool,
) -> None:
    records = tuple(
        make_record(f"s-{index}", amount, at(index + 2, 10))
        for index, amount in enumerate(pnl)
    )
    result = make_result(actual=records, shadow=())
    factor = build_report(result, context()).performance.net_profit_factor
    assert (factor.value, factor.is_unbounded, factor.is_undefined) == (value, unbounded, undefined)


def test_even_median_cagr_equity_drawdown_and_daily_zero_dates() -> None:
    records = (
        make_record("a", Decimal("500"), at(2, 10)),
        make_record("b", Decimal("-200"), at(3, 10)),
    )
    report = build_report(make_result(actual=records, shadow=()), context())

    assert report.performance.median_net_return_per_trade == Decimal("0.003")
    assert report.performance.cagr is not None and report.performance.cagr > 0
    assert len(report.equity_curve) == 3
    assert report.equity_curve[-1].realized_capital == Decimal("100300")
    assert report.performance.maximum_realized_drawdown == Decimal("200")
    assert report.performance.maximum_realized_drawdown_pct == Decimal("200") / Decimal("100500")
    assert len(report.daily_performance) == 5
    assert report.daily_performance[0].actual_trade_count == 0
    assert report.daily_performance[0].net_pnl == 0
    assert report.daily_performance[-1].realized_end_capital == Decimal("100300")
    assert report.performance.actual_trades_per_day == Decimal("0.4")


def test_same_timestamp_realization_is_atomic() -> None:
    records = (
        make_record("a", Decimal("500"), at(2, 10)),
        make_record("b", Decimal("-200"), at(2, 10)),
    )
    report = build_report(make_result(actual=records, shadow=()), context())
    assert len(report.equity_curve) == 2
    assert report.equity_curve[1].group_net_pnl == Decimal("300")
    assert report.equity_curve[1].drawdown_amount == 0


def test_zero_trade_report_has_defined_empty_semantics() -> None:
    report = build_report(
        make_result(actual=(), shadow=(), ending_capital=Decimal("100000")), context()
    )
    metrics = report.performance
    assert metrics.actual_trade_count == 0
    assert metrics.win_rate is None
    assert metrics.average_net_pnl_per_trade is None
    assert metrics.net_profit_factor.is_undefined
    assert metrics.cagr == 0
    assert metrics.actual_trades_per_day == 0
    assert len(report.equity_curve) == 1
    assert len(report.daily_performance) == 5


def test_breakdowns_are_sorted_separate_and_costs_are_retained() -> None:
    report = build_report(make_result(), context())
    strategies = [
        (row.strategy_id, row.strategy_version)
        for row in report.actual_strategy_breakdown
    ]
    assert strategies == [
        ("alpha", "1"),
        ("beta", "1"),
    ]
    assert [row.symbol for row in report.symbol_breakdown] == ["AAA", "BBB"]
    assert [row.exit_reason for row in report.actual_exit_reason_breakdown] == list(ExitReason)
    assert report.actual_costs.brokerage == Decimal("3")
    assert report.shadow_costs.brokerage == Decimal("1")
    assert report.actual_costs.component_percentages["brokerage"] == (
        Decimal("3") / Decimal("52.5")
    )
    assert not hasattr(report.actual_strategy_breakdown[0], "cagr")


def test_strict_acceptance_thresholds_and_frequency_separation() -> None:
    report = build_report(make_result(), context())
    assert not report.acceptance.average_net_return_pass
    assert not report.acceptance.hard_quantitative_targets_pass
    assert not hasattr(report.acceptance, "strategy_approved")

    exactly_ten = tuple(
        make_record(f"s-{index}", Decimal("250"), at(2 + index // 3, 10 + index % 3))
        for index in range(10)
    )
    frequent = build_report(make_result(actual=exactly_ten, shadow=()), context())
    assert frequent.performance.actual_trades_per_day == 2
    assert frequent.acceptance.frequency_target_met


def test_request_funnel_counts_every_outcome_without_misclassifying_exhaustion() -> None:
    result = make_result()
    actual_record = result.actual_trade_records[0]
    shadow_record = result.shadow_trade_records[0]

    def completed(
        record: BacktestTradeRecord,
        outcome: BacktestRequestOutcome,
    ) -> BacktestRequestResult:
        return BacktestRequestResult(
            request=BacktestTradeRequest(candidate=record.allocation_decision.candidate),
            outcome=outcome,
            terminal_at=record.trade.exit_fill.timestamp,
            allocation_decision=record.allocation_decision,
            trade_record=record,
        )

    actual_request = BacktestTradeRequest(candidate=actual_record.allocation_decision.candidate)
    shadow_request = BacktestTradeRequest(candidate=shadow_record.allocation_decision.candidate)
    request_results = (
        completed(actual_record, BacktestRequestOutcome.COMPLETED_ACTUAL),
        completed(shadow_record, BacktestRequestOutcome.COMPLETED_SHADOW),
        BacktestRequestResult(
            request=actual_request,
            outcome=BacktestRequestOutcome.ALLOCATED_ENTRY_NOT_FILLED,
            terminal_at=actual_request.candidate.order_intent.timestamp,
            allocation_decision=actual_record.allocation_decision,
        ),
        BacktestRequestResult(
            request=shadow_request,
            outcome=BacktestRequestOutcome.SHADOW_ENTRY_NOT_FILLED,
            terminal_at=shadow_request.candidate.order_intent.timestamp,
            allocation_decision=shadow_record.allocation_decision,
        ),
        BacktestRequestResult(
            request=actual_request,
            outcome=BacktestRequestOutcome.CAPITAL_EXHAUSTED,
            terminal_at=actual_request.candidate.order_intent.timestamp,
        ),
    )
    summary = build_report(
        result.model_copy(update={"request_results": request_results}), context()
    ).request_outcomes

    assert summary.total_requests == 5
    assert summary.completed_actual == 1
    assert summary.completed_shadow == 1
    assert summary.allocated_entry_not_filled == 1
    assert summary.shadow_entry_not_filled == 1
    assert summary.capital_exhausted == 1
    assert summary.capacity_rejected_request_count == 2
    assert summary.actual_completion_rate == Decimal("0.2")
    assert summary.capacity_rejection_rate == Decimal("0.4")
    assert summary.no_fill_rate == Decimal("0.4")


def test_oos_provenance_is_verified_without_mutation() -> None:
    result = make_result()
    fingerprint = fingerprint_backtest_result(result)
    record = OOSTestRecord(
        research_scope_id="scope",
        plan_id="plan",
        window_id="window",
        backtest_run_id=result.run_id,
        backtest_git_commit=result.git_commit,
        backtester_version=result.backtester_version,
        backtest_window_start=result.window_start,
        backtest_window_end=result.window_end,
        cost_policy_id=result.cost_policy_id,
        brokerage_plan=result.brokerage_plan.value,
        symbols=result.symbols,
        strategy_versions=result.strategy_versions,
        ml_model_versions=result.ml_model_versions,
        result_fingerprint=fingerprint,
        scope_strategy_ids=tuple(strategy_id for strategy_id, _ in result.strategy_versions),
        tested_strategy_versions=result.strategy_versions,
        registration_audit=OOSAuditContext(
            event_id="event",
            occurred_at=context().generated_at,
            git_commit="audit-commit",
        ),
    )
    before = record.model_dump()
    report = build_report(result, context(), record)
    assert report.provenance.oos_result_fingerprint == fingerprint
    assert record.model_dump() == before

    with pytest.raises(ReportingIntegrityError, match="run_id"):
        build_report(result, context(), record.model_copy(update={"backtest_run_id": "wrong"}))
    with pytest.raises(ReportingIntegrityError, match="fingerprint"):
        build_report(result, context(), record.model_copy(update={"result_fingerprint": "wrong"}))


def test_polars_tables_and_parquet_dataset_are_canonical_and_non_overwriting(
    tmp_path: Path,
) -> None:
    report = build_report(make_result(), context())
    tables = report_tables(report)
    assert tuple(tables) == tuple(REPORT_TABLE_FILENAMES)
    assert all(isinstance(table, pl.DataFrame) for table in tables.values())
    assert tables["actual_trades"].height == 3
    assert tables["shadow_trades"].height == 1
    assert tables["shadow_trades"]["economic_status"][0].startswith("SHADOW")
    assert report_tables(report)["actual_trades"].equals(tables["actual_trades"])

    paths = write_report_dataset(report, tmp_path / "dataset")
    assert {path.name for path in paths} == set(REPORT_TABLE_FILENAMES.values())
    summary = pl.read_parquet(tmp_path / "dataset" / "summary.parquet")
    assert summary["net_profit"][0] == Decimal("300")
    with pytest.raises(FileExistsError):
        write_report_dataset(report, tmp_path / "dataset")


def test_excel_uses_bundle_values_and_has_native_charts(tmp_path: Path) -> None:
    report = build_report(make_result(), context())
    path = write_excel_report(report, tmp_path / "report.xlsx")
    workbook = load_workbook(path, data_only=False)
    required = {
        "Dashboard",
        "Summary",
        "Actual Trades",
        "Shadow Trades",
        "Daily Performance",
        "Requests",
        "Equity Curve",
        "Strategy Breakdown",
        "Symbol Breakdown",
        "Costs",
        "Exit Reasons",
        "Provenance",
        "Cumulative PnL",
        "Monthly Performance",
        "Side Performance",
        "Time of Day",
        "Holding Distribution",
        "Rolling 20 Trades",
        "Trade Diagnostics",
        "Cost Impact",
        "Outcome Funnel",
        "Actual vs Shadow",
    }
    assert workbook.sheetnames[0] == "Dashboard"
    assert required <= set(workbook.sheetnames)
    assert workbook["Dashboard"]["B6"].value == 300
    assert workbook["Actual Trades"].max_row == 4
    assert workbook["Shadow Trades"].max_row == 2
    assert workbook["Daily Performance"].max_row == 6
    assert len(workbook["Dashboard"]._charts) >= 8
    assert len(workbook["Exit Reasons"]._charts) == 1
    assert len(workbook["Symbol Breakdown"]._charts) == 1


def test_matplotlib_visuals_are_headless_nonempty_and_close_figures(tmp_path: Path) -> None:
    report = build_report(make_result(), context())
    before = tuple(plt.get_fignums())
    paths = write_visual_report(report, tmp_path / "visuals")
    assert {path.name for path in paths} >= {
        "realized_equity_curve.png",
        "realized_drawdown.png",
        "daily_net_pnl.png",
        "trade_net_returns.png",
        "strategy_net_pnl.png",
        "cost_composition.png",
        "request_outcomes.png",
        "exit_reason_counts.png",
        "cumulative_gross_vs_net_pnl.png",
        "cumulative_cost_drag.png",
        "monthly_net_pnl.png",
        "monthly_trade_count.png",
        "monthly_win_rate.png",
        "monthly_profit_factor.png",
        "long_short_performance.png",
        "time_of_day_performance.png",
        "holding_time_distribution.png",
        "mfe_mae_scatter.png",
        "mfe_vs_realized_return.png",
        "mae_vs_realized_return.png",
        "rolling_win_rate.png",
        "rolling_profit_factor.png",
        "rolling_average_net_return.png",
        "symbol_performance_top_bottom.png",
        "actual_vs_shadow_comparison.png",
    }
    assert all(path.stat().st_size > 0 for path in paths)
    assert tuple(plt.get_fignums()) == before


def test_visual_report_preflights_entire_batch_before_writing(tmp_path: Path) -> None:
    report = build_report(make_result(), context())
    directory = tmp_path / "visuals"
    directory.mkdir()
    collision = directory / "cost_composition.png"
    collision.write_bytes(b"keep")
    with pytest.raises(FileExistsError, match="already exists"):
        write_visual_report(report, directory)
    assert collision.read_bytes() == b"keep"
    assert tuple(directory.iterdir()) == (collision,)


def test_zero_trade_excel_and_visuals_do_not_fabricate_observations(tmp_path: Path) -> None:
    report = build_report(
        make_result(actual=(), shadow=(), ending_capital=Decimal("100000")), context()
    )
    assert write_excel_report(report, tmp_path / "zero.xlsx").exists()
    paths = write_visual_report(report, tmp_path / "zero-visuals")
    assert len(paths) == 25
    assert all(path.stat().st_size > 0 for path in paths)


def test_visual_diagnostic_tables_are_deterministic_and_reconcile_exactly() -> None:
    report = build_report(make_result(), context())
    tables = report_tables(report)

    cumulative = tables["cumulative_pnl"].to_dicts()
    assert cumulative[-1]["cumulative_net_pnl"] == Decimal("300")
    assert cumulative[-1]["cumulative_cost_drag"] == report.actual_costs.total_costs
    assert (
        cumulative[-1]["cumulative_gross_pnl"]
        - cumulative[-1]["cumulative_cost_drag"]
        == cumulative[-1]["cumulative_net_pnl"]
    )
    monthly = tables["monthly_performance"].to_dicts()
    assert monthly == [
        {
            "month": "2025-01",
            "trade_count": 3,
            "net_pnl": Decimal("300"),
            "win_rate": Decimal("1") / Decimal("3"),
            "average_net_return": Decimal("0.002"),
            "total_costs": Decimal("52.5"),
            "profit_factor": Decimal("2.5"),
            "profit_factor_is_unbounded": False,
            "profit_factor_is_undefined": False,
        }
    ]
    assert tables["side_performance"].select(
        "side", "trade_count", "net_pnl"
    ).to_dicts() == [
        {"side": "LONG", "trade_count": 2, "net_pnl": Decimal("500")},
        {"side": "SHORT", "trade_count": 1, "net_pnl": Decimal("-200")},
    ]
    assert tables["time_of_day_performance"]["entry_time_bucket_ist"].to_list() == [
        "09:00"
    ]
    holding = tables["holding_time_distribution"].filter(
        pl.col("bucket") == ">30-60"
    )
    assert holding["trade_count"][0] == 3
    diagnostics = tables["trade_diagnostics"]
    assert diagnostics["holding_minutes"].to_list() == [Decimal("50")] * 3
    assert diagnostics["mfe_return"].to_list() == [Decimal("0.02")] * 3
    assert diagnostics["mae_return"].to_list() == [Decimal("-0.01")] * 3
    assert tables["rolling_trade_metrics"].is_empty()
    assert tables["actual_shadow_comparison"]["economic_status"][1].startswith(
        "HYPOTHETICAL"
    )


def test_rolling_metrics_begin_only_at_twenty_completed_actual_trades() -> None:
    records = tuple(
        make_record(
            f"s-{index}",
            Decimal("100") if index % 2 == 0 else Decimal("-50"),
            at(2, 10) + timedelta(minutes=index),
        )
        for index in range(20)
    )
    report = build_report(make_result(actual=records, shadow=()), context())
    rolling = report_tables(report)["rolling_trade_metrics"]

    assert rolling.height == 1
    assert rolling["trade_number"][0] == 20
    assert rolling["rolling_win_rate"][0] == Decimal("0.5")
    assert rolling["profit_factor"][0] == Decimal("2")


def _comparison_report(
    report_id: str,
    run_id: str,
    window_start: datetime,
    window_end: datetime,
    *,
    research_scope_id: str | None = None,
    plan_id: str | None = None,
    window_id: str | None = None,
):
    report = build_report(make_result(), context())
    return report.model_copy(
        update={
            "provenance": report.provenance.model_copy(
                update={
                    "report_id": report_id,
                    "run_id": run_id,
                    "window_start": window_start,
                    "window_end": window_end,
                    "research_scope_id": research_scope_id,
                    "plan_id": plan_id,
                    "window_id": window_id,
                }
            )
        }
    )


def test_development_comparison_is_chronological_and_has_no_pooled_cagr() -> None:
    first = _comparison_report("report-1", "run-1", at(1, 9), at(2, 15))
    second = _comparison_report("report-2", "run-2", at(3, 9), at(4, 15))

    comparison = build_report_comparison((second, first))

    assert not comparison.is_oos
    assert [row.report_id for row in comparison.rows] == ["report-1", "report-2"]
    assert not hasattr(comparison, "cagr")
    assert all(row.cagr == first.performance.cagr for row in comparison.rows)


def test_oos_comparison_validates_scope_plan_and_unique_windows() -> None:
    first = _comparison_report(
        "report-1",
        "run-1",
        at(1, 9),
        at(2, 15),
        research_scope_id="scope",
        plan_id="plan",
        window_id="window-1",
    )
    second = _comparison_report(
        "report-2",
        "run-2",
        at(3, 9),
        at(4, 15),
        research_scope_id="scope",
        plan_id="plan",
        window_id="window-2",
    )
    comparison = build_report_comparison((second, first))
    assert comparison.is_oos
    assert comparison.research_scope_id == "scope"
    assert [row.window_id for row in comparison.rows] == ["window-1", "window-2"]

    with pytest.raises(ValueError, match="research_scope_id"):
        build_report_comparison(
            (
                first,
                second.model_copy(
                    update={
                        "provenance": second.provenance.model_copy(
                            update={"research_scope_id": "other"}
                        )
                    }
                ),
            )
        )
    with pytest.raises(ValueError, match="plan_id"):
        build_report_comparison(
            (
                first,
                second.model_copy(
                    update={
                        "provenance": second.provenance.model_copy(
                            update={"plan_id": "other"}
                        )
                    }
                ),
            )
        )
    with pytest.raises(ValueError, match="unique"):
        build_report_comparison(
            (
                first,
                second.model_copy(
                    update={
                        "provenance": second.provenance.model_copy(
                            update={"window_id": "window-1"}
                        )
                    }
                ),
            )
        )


def test_comparison_outputs_are_auditable_headless_and_non_overwriting(
    tmp_path: Path,
) -> None:
    report = _comparison_report("report-1", "run-1", at(1, 9), at(2, 15))
    comparison = build_report_comparison((report,))
    excel = write_comparison_excel_report(comparison, tmp_path / "comparison.xlsx")
    workbook = load_workbook(excel)

    assert workbook.sheetnames == ["Comparison Dashboard", "Raw Comparison"]
    assert len(workbook["Comparison Dashboard"]._charts) == 8
    assert workbook["Raw Comparison"].max_row == 2
    with pytest.raises(FileExistsError):
        write_comparison_excel_report(comparison, excel)

    before = tuple(plt.get_fignums())
    paths = write_comparison_visual_report(comparison, tmp_path / "comparison-visuals")
    assert {path.name for path in paths} == set(COMPARISON_VISUAL_FILENAMES)
    assert all(path.stat().st_size > 0 for path in paths)
    assert tuple(plt.get_fignums()) == before
    with pytest.raises(FileExistsError):
        write_comparison_visual_report(comparison, tmp_path / "comparison-visuals")


def test_empty_and_partial_oos_comparisons_fail_cleanly() -> None:
    with pytest.raises(ValueError, match="at least one"):
        build_report_comparison(())

    development = _comparison_report("report-1", "run-1", at(1, 9), at(2, 15))
    partial = development.model_copy(
        update={
            "provenance": development.provenance.model_copy(
                update={"research_scope_id": "scope"}
            )
        }
    )
    with pytest.raises(ValueError, match="partial OOS"):
        build_report_comparison((partial,))
